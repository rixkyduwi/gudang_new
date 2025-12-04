# Import library bawaan Python
import io
import os
import textwrap
import locale
import uuid
import calendar
import time
from datetime import datetime, date, timedelta
from typing import Tuple, List, Optional
from decimal import Decimal, InvalidOperation
from collections import defaultdict

# Import library pihak ketiga
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from flask import (
    render_template,
    request,
    jsonify,
    g,
    send_file,
    Response,
    current_app,
)
from PIL import Image
from dateutil.relativedelta import relativedelta
from flask_jwt_extended import jwt_required, verify_jwt_in_request, get_jwt_identity
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfbase.pdfmetrics import stringWidth
from num2words import num2words

# Import dari aplikasi lokal
from . import app, mysql, render_pjax

# Middleware untuk membuka koneksi database sebelum setiap request
@app.before_request
def before_request():
    g.con = mysql.connection.cursor()
    g.con.execute("SET SESSION sql_mode=(SELECT REPLACE(@@sql_mode,'ONLY_FULL_GROUP_BY',''))")
    g.con.execute("SET @@sql_mode='';")
    g.start_time = time.time()

@app.after_request
def after_request(response):
    if hasattr(g, 'start_time'):
        elapsed = time.time() - g.start_time
        app.logger.info("%s %s selesai dalam %.2f detik",
                        request.method, request.path, elapsed)
    return response

# Middleware untuk menutup koneksi database setelah setiap request
@app.teardown_request
def teardown_request(exception):
    if hasattr(g, 'con'):
        g.con.close()

# Fungsi untuk mengelola gambar (upload, edit, delete)
def do_image(do, table, id):
    try:
        if do == "delete":
            filename = get_image_filename(table, id)
            delete_image(filename)
            return True

        # Upload gambar baru
        file = request.files['gambar']
        if file is None or file.filename == '':
            return "default.jpg"
        else:
            filename = get_image_filename(table, id)
            delete_image(filename)
            return resize_and_save_image(file, table, id)

    except KeyError:
        # Jika kunci 'gambar' tidak ada dalam request.files
        if do == "edit" and table == "galeri":
            return True
        reset = request.form.get('reset', 'false')
        if reset == "true":
            g.con.execute(f"UPDATE {table} SET gambar = %s WHERE id = %s", ("default.jpg", id))
            g.con.connection.commit()
        return "default.jpg"

    except FileNotFoundError:
        pass  # Jika file tidak ditemukan, abaikan

    except Exception as e:
        print(str(e))
        return str(e)

# Fungsi untuk mengubah ukuran dan menyimpan gambar
def resize_and_save_image(file, table=None, id=None):
    img = Image.open(file).convert('RGB').resize((600, 300))
    random_name = uuid.uuid4().hex + ".jpg"
    destination = os.path.join(app.config['UPLOAD_FOLDER'], random_name)
    img.save(destination)

    if table and id:
        g.con.execute(f"UPDATE {table} SET gambar = %s WHERE id = %s", (random_name, id))
        g.con.connection.commit()
        return True
    else:
        return random_name

# Fungsi untuk mendapatkan nama file gambar dari database
def get_image_filename(table, id):
    g.con.execute(f"SELECT gambar FROM {table} WHERE id = %s", (id,))
    result = g.con.fetchone()
    if result == "default.jpg":
        return None
    return result[0] if result else None

# Fungsi untuk menghapus file gambar dari server
def delete_image(filename):
    if filename == "default.jpg":
        return True
    if filename:
        image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(image_path):
            os.remove(image_path)

# Fungsi untuk mengambil data dari database dan mengubahnya menjadi format dictionary
def fetch(query, params=None):
    print(query)
    print(params)
    g.con.execute(query, params or ())
    data = g.con.fetchall()
    column_names = [desc[0] for desc in g.con.description]
    return [dict(zip(column_names, row)) for row in data]

# Fungsi untuk mengambil daftar tahun dari database
def fetch_years(pilih):
    if pilih == "barang_masuk":
        query = "SELECT YEAR(tglfaktur) AS tahun FROM barang_masuk GROUP BY tahun"
    elif pilih == "barang_keluar":
        query = "SELECT YEAR(tglfaktur) AS tahun FROM barang_keluar GROUP BY tahun"
    else:
        query = ""
    g.con.execute(query)
    data_thn = g.con.fetchall()
    return [{'tahun': str(sistem[0])} for sistem in data_thn]

def fetch_sum(query, params=None):
    g.con.execute(query, params or ())
    row = g.con.fetchone()
    return row[0] if row and row[0] is not None else 0
# form json    
def req(key):
    return request.json.get(key)

# update stok gudang
def update_stok_gudang(id_barang, jmlpermintaan, operasi):
    jml = int(jmlpermintaan)

    # Ambil stok gudang dan limit sekali query
    g.con.execute("""
        SELECT bg.sisa_gudang, b.stoklimit
        FROM barang b
        LEFT JOIN barang_gudang bg ON bg.id_barang = b.id
        WHERE b.id = %s
    """, (id_barang,))
    row = g.con.fetchone()
    
    print("=" * 40)
    print(f"DEBUG update_stok_gudang")
    print(f"id_barang     : {id_barang}")
    print(f"jmlpermintaan : {jmlpermintaan}")
    print(f"operasi       : {operasi}")
    print(f"sisa_gudang   : {row[0] if row else 'None'}")
    print(f"stoklimit     : {row[1] if row else 'None'}")

    if not row:
        return jsonify({"error": "Data barang tidak ditemukan"}), 404

    sisa_gudang, stoklimit = row if row[0] is not None else (0, row[1])

    if operasi == "kurangi":
        new_qty = sisa_gudang - jml
    else:
        new_qty = sisa_gudang + jml

    if new_qty < 0:
        return jsonify({"error": "Stok Kurang"}), 404

    ket = "Stok Aman" if new_qty > stoklimit else "Stok Tidak Aman"

    print(f"keterangan    : {ket}")
    print("=" * 40)
    # Insert jika belum ada di barang_gudang
    g.con.execute("""
        INSERT INTO barang_gudang (id_barang, sisa_gudang, keterangan)
        VALUES (%s, %s, %s)
        ON DUPLICATE KEY UPDATE
            sisa_gudang = VALUES(sisa_gudang),
            keterangan = VALUES(keterangan)
    """, (id_barang, new_qty, ket))
    g.con.connection.commit()

    return True

#function di render_template
@app.template_filter('format_currency')
def format_currency(value):
    locale.setlocale(locale.LC_ALL, 'id_ID.UTF-8')
    return locale.currency(value, grouping=True, symbol='Rp')

@app.template_filter('clean_currency')
def clean_currency(value):
    """Remove 'Rp', dots, and spaces, then convert to float"""
    cleaned_value = value.replace('Rp', '').replace('RP', '').replace('.', '').replace(',', '.').strip()
    return float(cleaned_value)

@app.template_filter('floatformat')
def floatformat(value, precision=2):
    try:
        return f"{float(value):.{precision}f}"
    except (ValueError, TypeError):
        return value

@app.template_filter('format_rp')
def format_rp(value):
    value =f"{int(value):,}".replace(",", ".")
    value = f"Rp. {value},00"
    return value

@app.template_filter('format_rupiah')
def format_rupiah(value):
    try:
        # Ubah ke Decimal dan pastikan dua digit di belakang koma
        value = Decimal(value).quantize(Decimal("0.01"))

        # Pecah jadi bagian depan dan koma
        bagian_utama, bagian_koma = str(value).split('.')

        # Format bagian utama dengan titik pemisah ribuan
        bagian_utama = f"{int(bagian_utama):,}".replace(",", ".")

        # Gabungkan dan pakai koma sebagai desimal (gaya Indonesia)
        return f"{bagian_utama},{bagian_koma}"
    except (InvalidOperation, ValueError, TypeError):
        return value

@app.template_filter('format_date')
def format_date(value):
    if not value:
        return ""

    # Kalau sudah datetime/date object
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")

    # Kalau string, coba parse dengan beberapa format
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                dt = datetime.strptime(value, fmt)
                return dt.strftime("%d/%m/%Y")
            except ValueError:
                continue

    # Fallback kalau gak dikenali
    return str(value)

list_bulan = [{"value":"1","nama_bulan":"Januari"},{"value":"2","nama_bulan":"Februari"},{"value":"3","nama_bulan":"Maret"},
                {"value":"4","nama_bulan":"April"},{"value":"5","nama_bulan":"Mei"},{"value":"6","nama_bulan":"Juni"},{"value":"7","nama_bulan":"Juli"},
                {"value":"8","nama_bulan":"Agustus"},{"value":"9","nama_bulan":"September"},{"value":"10","nama_bulan":"Oktober"},
                {"value":"11","nama_bulan":"November"},{"value":"12","nama_bulan":"Desember"}] 

def time_zone_wib():
    return datetime.utcnow() + timedelta(hours=7)
@app.route('/admin/dashboard')
@jwt_required()
def dashboard():
    user = get_jwt_identity()
    # --- tanggal & rentang waktu (WIB) ---
    today = time_zone_wib()
    hari_ini   = today.strftime("%d/%m/%Y")
    tanggal_db = today.strftime("%Y-%m-%d")

    bulan_ini = today.month
    tahun_ini = today.year

    # rentang bulan berjalan [start_month, end_month)
    last_day     = calendar.monthrange(tahun_ini, bulan_ini)[1]
    start_month  = date(tahun_ini, bulan_ini, 1)
    end_month    = date(tahun_ini, bulan_ini, last_day) + timedelta(days=1)

    # rentang hari ini [start_day, end_day)
    start_day = date.fromisoformat(tanggal_db)
    end_day   = start_day + timedelta(days=1)

    # -------------------------
    # Target sales + realtime sales bulan ini per salesperson
    # -------------------------
    target_sales = fetch("""
        SELECT 
            t.id,
            s.name AS nama_sales,
            t.target_amount AS target,
            s.id AS id_sales,
            COALESCE(m.total_amount, 0) AS data_realtime
        FROM sales_targets t
        JOIN salespersons s ON s.id = t.salesperson_id
        LEFT JOIN (
            SELECT si.salesperson_id, SUM(sit.total_amount) AS total_amount
            FROM sales_invoices si
            JOIN sales_items    sit ON sit.sales_invoice_id = si.id
            WHERE si.invoice_date >= %s AND si.invoice_date < %s
            GROUP BY si.salesperson_id
        ) m ON m.salesperson_id = s.id
        WHERE t.year = %s AND t.month = %s
        ORDER BY s.name ASC
    """, (start_month, end_month, tahun_ini, bulan_ini))

    # -------------------------
    # Inkaso (penerimaan pembayaran hari ini) – ambil dari payments
    # -------------------------
    g.con.execute("""
        SELECT COALESCE(SUM(p.amount), 0)
        FROM payments p
        WHERE p.ref_type = 'SALE'
          AND p.pay_date >= %s AND p.pay_date < %s
    """, (start_day, end_day))
    inkaso = g.con.fetchone()[0] or 0

    # -------------------------
    # Performa penjualan bulan ini – total dari sales_items
    # -------------------------
    g.con.execute("""
        SELECT COALESCE(SUM(sit.total_amount), 0)
        FROM sales_items sit
        JOIN sales_invoices si ON si.id = sit.sales_invoice_id
        WHERE si.invoice_date >= %s AND si.invoice_date < %s
    """, (start_month, end_month))
    performa_penjualan = g.con.fetchone()[0] or 0

    # -------------------------
    # Performa belanja bulan ini – total dari purchase_items
    # -------------------------
    g.con.execute("""
        SELECT COALESCE(SUM(pi.total_amount), 0)
        FROM purchase_items pi
        JOIN purchases p ON p.id = pi.purchase_id
        WHERE p.invoice_date >= %s AND p.invoice_date < %s
    """, (start_month, end_month))
    performa_belanja = g.con.fetchone()[0] or 0

    revenue = (performa_penjualan or 0) - (performa_belanja or 0)

    return render_pjax(
        'admin/dashboard.html',
        tahun_ini=tahun_ini,
        target_sales=target_sales,
        inkaso=inkaso,
        hari_ini=hari_ini,
        revenue=revenue,
        performa_penjualan=performa_penjualan,
        performa_belanja=performa_belanja
    )

def one(sql, params=()):
    g.con.execute(sql, params)
    return g.con.fetchone()

def parse_decimal(x, default=Decimal('0')):
    try:
        return Decimal(str(x))
    except (InvalidOperation, TypeError, ValueError):
        return default

@app.route('/admin/penerimaan-tambah', methods=['GET'])
def admin_penerimaan_tambah():
    data_barang = fetch("SELECT id, code, name, unit, stock_min FROM products WHERE is_active=1 ORDER BY name")
    data_supplier = fetch("SELECT id, name, address, phone FROM suppliers ORDER BY name")
    return render_pjax(
        "admin/tambah-penerimaan.html",
        data_supplier=data_supplier,
        data_barang=data_barang,
        tanggal=time_zone_wib().date()
    )

@app.route('/admin/penerimaan')
def admin_penerimaan():
    tahun   = request.args.get('tahun', type=int)
    bulan   = request.args.get('bulan', type=int)
    tanggal = request.args.get('tanggal', type=int)
    supplier_name = request.args.get('nama_principle', type=str)
    page    = request.args.get('page', 1, type=int)
    per_page= request.args.get('per_page', 10, type=int)

    filters, params = [], []
    clause, prms = build_date_range(year=tahun, month=bulan, day=tanggal, alias="p", col="invoice_date")
    if clause:
        filters.append(clause); params += prms
    if supplier_name:
        filters.append("s.name = %s"); params.append(supplier_name)

    where = "WHERE " + " AND ".join(filters) if filters else ""

    # Hitung total distinct purchases
    cnt_sql = f"""
        SELECT COUNT(*) FROM (
            SELECT p.id
            FROM purchases p
            JOIN suppliers s ON s.id = p.supplier_id
            {where}
            GROUP BY p.id
        ) x
    """
    g.con.execute(cnt_sql, params)
    total_records = g.con.fetchone()[0] or 0

    offset = (page - 1) * per_page

    # Ambil header + total_amount (SUM pi.total_amount)
    rows = fetch(f"""
        SELECT p.id,
               p.invoice_date AS tglfaktur,
               p.invoice_no   AS nofaktur,
               s.name         AS nama_supplier,
               COALESCE(SUM(pi.total_amount),0) AS performa_belanja
        FROM purchases p
        JOIN suppliers s    ON s.id = p.supplier_id
        JOIN purchase_items pi ON pi.purchase_id = p.id
        {where}
        GROUP BY p.id, p.invoice_date, p.invoice_no, s.name
        ORDER BY p.id DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])

    info_list = rows or []
    if info_list:
        # Ambil semua detail untuk p.id yang tampil
        ids = tuple(r['id'] for r in info_list)
        data_detail = fetch(f"""
            SELECT pi.id,
                   pi.purchase_id,
                   pi.product_id,
                   pr.code AS kode_barang,
                   pr.name AS nama_barang,
                   pr.unit AS base_unit,
                   pi.qty AS jml_menerima, 
                   pi.unit_price AS harga_satuan, 
                   pi.total_amount AS harga_total,
                   -- UoM-lite (jika dipakai di form)
                   pi.unit_label, pi.uom_factor_to_base, pi.qty_uom, pi.qty_base, pi.unit_price_uom, pi.unit_price_base
            FROM purchase_items pi
            JOIN products pr ON pr.id = pi.product_id
            WHERE pi.purchase_id IN %s
            ORDER BY pi.id
        """, (ids,))
        # tempel ke header
        for h in info_list:
            h['detail'] = [d for d in data_detail if d['purchase_id'] == h['id']]

    # Pagination info
    total_pages = (total_records + per_page - 1) // per_page
    return render_pjax(
        "admin/penerimaan.html",
        info_list=info_list,
        data_supplier=fetch("SELECT id, name FROM suppliers ORDER BY name"),
        data_barang=fetch("SELECT id, code, name, unit FROM products WHERE is_active=1 ORDER BY name"),
        list_tahun=fetch("SELECT YEAR(invoice_date) as tahun from purchases group by tahun"),
        page=page, per_page=per_page,
        total_pages=total_pages, total_records=total_records,
        has_next=page < total_pages, has_prev=page > 1,
        page_range=range(max(1, page-3), min(total_pages+1, page+3)),
        tanggal_hari_ini=time_zone_wib().date()
    )

@app.route('/admin/penerimaan', methods=['POST'])
@jwt_required()
def tambah_penerimaan():
    d = request.get_json() or {}
    supplier_id = d.get('supplier_id') or d.get('id_supplier')
    invoice_no  = d.get('invoice_no')  or d.get('nofaktur')
    invoice_date= d.get('invoice_date')or d.get('tglfaktur')
    due_date    = d.get('due_date') or d.get('jthtempo')
    items       = d.get('items') or []

    if not supplier_id or not invoice_no or not invoice_date or not items:
        return jsonify({"error":"supplier_id, invoice_no, invoice_date, items wajib"}), 400

    # Validasi supplier
    if not one("SELECT id FROM suppliers WHERE id=%s", (supplier_id,)):
        return jsonify({"error":"Supplier not found"}), 404

    try:
        # Header
        g.con.execute("""
            INSERT INTO purchases (supplier_id, invoice_no, invoice_date, due_date, status)
            VALUES (%s,%s,%s,%s,'POSTED')
        """, (supplier_id, invoice_no, invoice_date, due_date))
        purchase_id = g.con.lastrowid

        # Detail
        total_hdr = Decimal('0.00')
        for it in items:
            product_id  = it.get('product_id') or it.get('id_barang')
            qty         = int(it.get('qty') or it.get('jml_menerima') or 0)
            unit_price  = parse_decimal(it.get('unit_price') or it.get('harga_satuan'))
            total_amount= parse_decimal(it.get('total_amount') or it.get('harga_total'))
            if total_amount != unit_price * qty:
                return jsonify({"error": f"Total item tidak sesuai, seharusnya {unit_price*qty}"}), 400

            # UoM-lite (opsional di form)
            unit_label = it.get('unit_label')
            factor     = parse_decimal(it.get('uom_factor_to_base') or (1 if not unit_label else 1))
            qty_uom    = parse_decimal(it.get('qty_uom') or (qty if not unit_label else 0))
            qty_base   = parse_decimal(it.get('qty_base') or (qty_uom*factor if unit_label else qty))
            unit_price_uom  = parse_decimal(it.get('unit_price_uom') or (unit_price if not unit_label else 0))
            unit_price_base = parse_decimal(it.get('unit_price_base') or ((unit_price_uom/factor) if unit_label and factor else unit_price))
            g.con.execute("""
                INSERT INTO purchase_items
                (purchase_id, product_id, qty, unit_price, total_amount,
                 unit_label, uom_factor_to_base, qty_uom, qty_base, unit_price_uom, unit_price_base)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (purchase_id, product_id, qty, unit_price, total_amount,
                  unit_label, factor, qty_uom, qty_base, unit_price_uom, unit_price_base))

            # Stock move (IN)
            g.con.execute("""
                INSERT INTO stock_moves (product_id, ref_type, ref_id, qty_in, note)
                VALUES (%s,'PURCHASE',%s,%s,%s)
            """, (product_id, purchase_id, int(qty_base), f"Invoice {invoice_no}"))

            total_hdr += total_amount

        g.con.connection.commit()
        return jsonify({"msg":"SUKSES","id":purchase_id,"total":str(total_hdr)})
    except Exception as e:
        g.con.connection.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/admin/penerimaan/<int:id>', methods=['GET'])
def penerimaan_detail(id):
    hdr = fetch("""
        SELECT p.id, p.supplier_id,  p.invoice_date, p.invoice_no,
                s.name AS supplier_name,
                s.address AS supplier_address,
                s.phone AS supplier_phone,
               p.status
        FROM purchases p
        JOIN suppliers s ON s.id = p.supplier_id
        WHERE p.id=%s
    """, (id,))
    if not hdr:
        return "Not found", 404
    items = fetch("""
        SELECT pi.id, pi.product_id, pr.code, pr.name, pr.unit,
               pi.qty, pi.unit_price, pi.total_amount,
               pi.unit_label, pi.uom_factor_to_base, pi.qty_uom, pi.qty_base, pi.unit_price_uom, pi.unit_price_base
        FROM purchase_items pi
        JOIN products pr ON pr.id = pi.product_id
        WHERE pi.purchase_id=%s
        ORDER BY pi.id
    """, (id,))
    data_supplier = fetch("SELECT id, name FROM suppliers ORDER BY name")
    data_barang   = fetch("SELECT id, code, name, unit FROM products WHERE is_active=1 ORDER BY name")
    return render_pjax("admin/edit_penerimaan.html",
                           purchase=hdr[0], items=items,
                           data_supplier=data_supplier, data_barang=data_barang,
                           tanggal=time_zone_wib().date())

@app.route('/admin/penerimaan/<int:id>', methods=['PUT'])
@jwt_required()
def penerimaan_update(id):
    d = request.get_json() or {}
    supplier_id = d.get('supplier_id')
    invoice_no  = d.get('invoice_no')
    invoice_date= d.get('invoice_date')
    items       = d.get('items') or []
    if not supplier_id or not invoice_no or not invoice_date or not items:
        return jsonify({"error":"supplier_id, invoice_no, invoice_date, items wajib"}), 400

    try:
        # Update header
        g.con.execute("""
            UPDATE purchases SET supplier_id=%s, invoice_no=%s, invoice_date=%s
            WHERE id=%s
        """, (supplier_id, invoice_no, invoice_date, id))

        # Hapus items & stock_moves lama
        g.con.execute("DELETE FROM stock_moves WHERE ref_type='PURCHASE' AND ref_id=%s", (id,))
        g.con.execute("DELETE FROM purchase_items WHERE purchase_id=%s", (id,))

        total_hdr = Decimal('0')
        for it in items:
            product_id  = it.get('product_id')
            qty         = int(it.get('qty') or it.get('jml_menerima') or 0)
            unit_price  = parse_decimal(it.get('unit_price') or it.get('harga_satuan'))
            total_amount= parse_decimal(it.get('total_amount') or it.get('harga_total'))
            if total_amount != unit_price * qty:
                return jsonify({"error": f"Total item tidak sesuai, seharusnya {unit_price*qty}"}), 400

            unit_label = it.get('unit_label')
            factor     = parse_decimal(it.get('uom_factor_to_base') or (1 if not unit_label else 1))
            qty_uom    = parse_decimal(it.get('qty_uom') or (qty if not unit_label else 0))
            qty_base   = parse_decimal(it.get('qty_base') or (qty_uom*factor if unit_label else qty))
            unit_price_uom  = parse_decimal(it.get('unit_price_uom') or (unit_price if not unit_label else 0))
            unit_price_base = parse_decimal(it.get('unit_price_base') or ((unit_price_uom/factor) if unit_label and factor else unit_price))

            g.con.execute("""
                INSERT INTO purchase_items
                (purchase_id, product_id, qty, unit_price, total_amount,
                 unit_label, uom_factor_to_base, qty_uom, qty_base, unit_price_uom, unit_price_base)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (id, product_id, qty, unit_price, total_amount,
                  unit_label, factor, qty_uom, qty_base, unit_price_uom, unit_price_base))

            g.con.execute("""
                INSERT INTO stock_moves (product_id, ref_type, ref_id, qty_in, note)
                VALUES (%s,'PURCHASE',%s,%s,%s)
            """, (product_id, id, int(qty_base), f"Edit Invoice {invoice_no}"))

            total_hdr += total_amount

        g.con.connection.commit()
        return jsonify({"msg":"SUKSES","total":str(total_hdr)})
    except Exception as e:
        g.con.connection.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/admin/penerimaan/<int:id>', methods=['DELETE'])
@jwt_required()
def penerimaan_delete(id):
    try:
        # hapus dulu stock_moves pembelian ini
        g.con.execute("DELETE FROM stock_moves WHERE ref_type='PURCHASE' AND ref_id=%s", (id,))
        # hapus detail
        g.con.execute("DELETE FROM purchase_items WHERE purchase_id=%s", (id,))
        # hapus header
        g.con.execute("DELETE FROM purchases WHERE id=%s", (id,))
        g.con.connection.commit()
        return jsonify({"msg":"BERHASIL DIHAPUS"})
    except Exception as e:
        g.con.connection.rollback()
        return jsonify({"error": str(e)}), 500
        
@app.route('/admin/penyimpanan')
def adminpenyimpanan():
    # Param lama tetap didukung biar nggak breaking
    product_id = request.args.get('product_id') or request.args.get('id_barang')
    q          = request.args.get('q')  # optional: cari kode/nama
    low_only   = request.args.get('low_only', default=0, type=int)  # 1 = hanya stok tidak aman
    page       = request.args.get('page', default=1, type=int)
    per_page   = request.args.get('per_page', default=10, type=int)

    # Base FROM: ambil dari view v_product_stock + join products buat meta (code, name, unit, stock_min)
    base_from = """
        FROM v_product_stock s
        JOIN products p ON p.id = s.product_id
    """

    # Filter dinamis
    filters = []
    params  = []
    if product_id:
        filters.append("s.product_id = %s")
        params.append(product_id)

    if q:
        filters.append("(p.code LIKE %s OR p.name LIKE %s)")
        like = f"%{q}%"
        params.extend([like, like])

    # low_only: stok on hand <= stock_min
    if low_only == 1:
        filters.append("(s.qty_on_hand <= p.stock_min)")

    where_clause = "WHERE " + " AND ".join(filters) if filters else ""

    # Count total row (untuk pagination)
    count_sql = f"SELECT COUNT(*) {base_from} {where_clause}"
    g.con.execute(count_sql, params)
    total_records = g.con.fetchone()[0]

    # Paging
    total_pages = (total_records + per_page - 1) // per_page
    offset = (page - 1) * per_page

    # Data list
    list_sql = f"""
        SELECT
          s.product_id,
          p.code       AS kode_barang,
          p.name       AS nama_barang,
          p.unit,
          p.stock_min  AS stok_limit,
          s.qty_on_hand AS sisa_gudang,
          CASE
            WHEN s.qty_on_hand <= p.stock_min THEN 'Stok Tidak Aman'
            ELSE 'Aman'
          END AS status
        {base_from}
        {where_clause}
        ORDER BY s.product_id ASC
        LIMIT %s OFFSET %s
    """
    list_params = params + [per_page, offset]
    info_list = fetch(list_sql, list_params)
    print(info_list)
    # Dropdown daftar barang (untuk filter select)
    nama_barang_list = fetch("""
        SELECT p.id, p.code AS kode_barang, p.name AS nama_barang
        FROM products p
        ORDER BY p.name ASC
    """)

    has_next = page < total_pages
    has_prev = page > 1
    page_range = range(max(1, page - 3), min(total_pages + 1, page + 3))

    return render_pjax(
        "admin/penyimpanan.html",
        info_list=info_list,
        nama_barang=nama_barang_list,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        total_records=total_records,
        has_next=has_next,
        has_prev=has_prev,
        page_range=page_range,
        q=q,
        product_id=product_id,
        low_only=low_only
    )
    
@app.route('/admin/penyimpanan', methods=['PUT'])
@jwt_required()
def edit_penyimpanan():
    """
    Payload JSON:
    {
      "product_id": 123,
      "new_qty": 50,         # stok final yang diinginkan
      "note": "Stock opname Sept"
    }
    """
    data = request.json or {}
    print(data)
    product_id = data.get('product_id')
    new_qty    = data.get('new_qty')
    note       = data.get('note') or 'Manual adjustment'

    if not product_id or new_qty is None:
        return jsonify({"error": "product_id dan new_qty wajib"}), 400

    try:
        # 1) Ambil stok saat ini (qty_on_hand) + batas aman (stock_min)
        sql_now = """
            SELECT s.qty_on_hand, p.stock_min, p.name
            FROM v_product_stock s
            JOIN products p ON p.id = s.product_id
            WHERE s.product_id = %s
        """
        g.con.execute(sql_now, (product_id,))
        row = g.con.fetchone()
        if not row:
            return jsonify({"error": "Produk tidak ditemukan"}), 404

        current_qty = int(row[0] or 0)
        stock_min   = int(row[1] or 0)
        product_name= row[2]

        # 2) Hitung delta
        new_qty = int(new_qty)
        delta = new_qty - current_qty
        if delta == 0:
            status = "Stok Tidak Aman" if new_qty <= stock_min else "Aman"
            return jsonify({
                "msg": "Tidak ada perubahan stok",
                "product_id": product_id,
                "product_name": product_name,
                "qty_on_hand": new_qty,
                "status": status
            })

        # 3) Insert stock_moves (ADJUSTMENT)
        if delta > 0:
            g.con.execute(
                "INSERT INTO stock_moves (product_id, ref_type, qty_in,  note) VALUES (%s,'ADJUSTMENT',%s,%s)",
                (product_id, delta, note)
            )
        else:
            g.con.execute(
                "INSERT INTO stock_moves (product_id, ref_type, qty_out, note) VALUES (%s,'ADJUSTMENT',%s,%s)",
                (product_id, abs(delta), note)
            )

        g.con.connection.commit()

        # 4) Ambil stok terbaru setelah insert (opsional)
        g.con.execute(sql_now, (product_id,))
        row2 = g.con.fetchone()
        new_now = int(row2[0] or 0)
        status = "Stok Tidak Aman" if new_now <= stock_min else "Aman"

        return jsonify({
            "msg": "SUKSES",
            "product_id": product_id,
            "product_name": product_name,
            "qty_on_hand": new_now,
            "status": status,
            "delta": delta
        })
    except Exception as e:
        g.con.connection.rollback()
        print(e)
        return jsonify({"error": str(e)}), 500

@app.route('/admin/penyimpanan/hapus/id', methods=['DELETE'])
@jwt_required()
def hapus_id_penyimpanan():
    try:
        id = request.json.get('id')
        # Hapus dari barang_gudang
        g.con.execute("DELETE FROM barang_gudang WHERE id = %s", (int(id),))
        g.con.connection.commit()

        return jsonify({"msg": "BERHASIL DIHAPUS"})
    except Exception as e:
        print(e)
        return jsonify({"error": str(e)}), 500

def next_invoice_no_for_date(d):
    row = one("""
        SELECT RIGHT(invoice_no, 3) AS last_3
        FROM sales_invoices
        WHERE invoice_date=%s
        ORDER BY invoice_no DESC
        LIMIT 1
    """, (d,))
    last = int(row[0]) if row and row[0] else 0
    return f"{d.strftime('%Y%m%d')}{last+1:03d}"


@app.route('/admin/pengeluaran-tambah')
def tambah_pengeluaran():
    # Produk + stok (gunakan view v_product_stock)
    data_barang = fetch("""
        SELECT p.id, p.code, p.name, p.unit, COALESCE(v.qty_on_hand,0) AS qty_on_hand, p.stock_min
        FROM products p
        LEFT JOIN v_product_stock v ON v.product_id = p.id
        WHERE p.is_active=1
        ORDER BY p.name
    """)
    # Sales & Customer
    data_sales = fetch("SELECT id, name FROM salespersons WHERE is_active=1 ORDER BY name")
    data_customers = fetch("SELECT id, name, address FROM customers ORDER BY name")

    # tanggal & jth tempo default
    today = time_zone_wib().date()
    jth_tempo = (today + timedelta(days=30)).strftime("%Y-%m-%d")
    nofaktur  = next_invoice_no_for_date(today)

    return render_pjax("admin/tambah-pengeluaran.html",
        data_sales=data_sales,
        data_customers=data_customers,
        data_barang=data_barang,
        tanggal=today.strftime("%Y-%m-%d"),
        nofaktur=nofaktur,
        jth_tempo=jth_tempo
    )

@app.route('/api/sales/invoice_no/<path:tanggal>')
def api_sales_invoice_no(tanggal):
    dt = None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(tanggal, fmt).date()
            break
        except ValueError:
            continue
    if not dt:
        return jsonify({"error":"Format tanggal tidak dikenali (YYYY-MM-DD / DD/MM/YYYY)"}), 400

    nofaktur = next_invoice_no_for_date(dt)
    jth_tempo = (dt + timedelta(days=30)).strftime("%d/%m/%Y")
    return jsonify({"nofaktur": nofaktur, "jatuh_tempo": jth_tempo})

@app.route('/admin/pengeluaran')
def admin_pengeluaran():
    tahun   = request.args.get('tahun', type=int)
    bulan   = request.args.get('bulan', type=int)
    tanggal = request.args.get('tanggal', type=int)
    salesperson = request.args.get('nama_sales', type=str)   # boleh ID atau nama
    customer    = request.args.get('nama_outlet', type=str)  # boleh ID atau nama
    page     = request.args.get('page', default=1, type=int)
    per_page = request.args.get('per_page', default=10, type=int)

    filters, params = [], []
    clause, prms = build_date_range(year=tahun, month=bulan, day=tanggal, alias="si", col="invoice_date")
    if clause: filters.append(clause); params += prms

    # filter optional
    if salesperson:
        # coba asumsikan nama; kalau kamu kirim ID, ganti ke s.id=%s
        filters.append("s.name = %s"); params.append(salesperson)
    if customer:
        filters.append("c.name = %s"); params.append(customer)

    where = "WHERE " + " AND ".join(filters) if filters else ""

    # hitung total distinct invoice
    cnt = one(f"""
        SELECT COUNT(*) FROM (
          SELECT si.id
          FROM sales_invoices si
          JOIN salespersons s ON s.id = si.salesperson_id
          LEFT JOIN customers c ON c.id = si.customer_id
          {where}
          GROUP BY si.id
        ) x
    """, tuple(params))
    total_records = cnt[0] if cnt else 0

    offset = (page-1)*per_page
    rows = fetch(f"""
        SELECT si.id,
               si.invoice_no,
               si.invoice_date,
               si.due_date,
               si.status,
               s.name AS nama_sales,
               COALESCE(c.name,'-') AS nama_customer,
               COALESCE(SUM(si2.total_amount),0) AS performa_sales
        FROM sales_invoices si
        JOIN salespersons s ON s.id = si.salesperson_id
        LEFT JOIN customers c ON c.id = si.customer_id
        LEFT JOIN sales_items si2 ON si2.sales_invoice_id = si.id
        {where}
        GROUP BY si.id, si.invoice_no, si.invoice_date, si.due_date, si.status, s.name, c.name
        ORDER BY si.id DESC
        LIMIT %s OFFSET %s
    """, tuple(params+[per_page, offset]))

    # detail untuk halaman ini
    detail_map = {}
    if rows:
        ids = tuple(r['id'] for r in rows)
        details = fetch("""
            SELECT it.sales_invoice_id,
                   it.product_id, p.code, p.name,
                   it.qty, it.unit_price, it.discount_percent, it.total_amount,
                   it.unit_label, it.uom_factor_to_base, it.qty_uom, it.qty_base
            FROM sales_items it
            JOIN products p ON p.id = it.product_id
            WHERE it.sales_invoice_id IN %s
            ORDER BY it.id
        """, (ids,))
        for d in details:
            detail_map.setdefault(d['sales_invoice_id'], []).append(d)
        for r in rows: r['detail'] = detail_map.get(r['id'], [])

    # dropdown filter
    nama_sales = fetch("SELECT name FROM salespersons WHERE is_active=1 ORDER BY name")
    nama_customer = fetch("SELECT name FROM customers ORDER BY name")

    total_pages = (total_records + per_page - 1) // per_page
    return render_pjax("admin/pengeluaran.html",
        info_list=rows,
        tahun=tahun, bulan=bulan, tanggal=tanggal,
        nama_sales=nama_sales, nama_customer=nama_customer,
        page=page, per_page=per_page, total_pages=total_pages, total_records=total_records,
        has_next=page < total_pages, has_prev=page > 1,
        page_range=range(max(1, page-3), min(total_pages+1, page+3))
    )

@app.route('/admin/pengeluaran', methods=['POST'])
@jwt_required()
def tambah_pengeluaran_action():
    d = request.get_json() or {}
    invoice_no   = d.get('nofaktur') or d.get('invoice_no')
    invoice_date = d.get('tglfaktur') or d.get('invoice_date')
    due_date     = d.get('jthtempo') or d.get('due_date')  # format YYYY-MM-DD
    pay_term     = d.get('pembayaran')  # "CASH"/"TEMPO" → mapping opsional
    tax_flag     = d.get('pajak')       # "Yes"/"No" → optional
    salesperson_name = d.get('nama_sales')
    customer_name    = d.get('nama_outlet')  # optional
    items = d.get('items') or []

    if not (invoice_no and invoice_date and salesperson_name and items):
        return jsonify({"error":"nofaktur, tglfaktur, nama_sales, items wajib"}), 400

    # resolve salesperson & customer
    sp = one("SELECT id FROM salespersons WHERE name=%s", (salesperson_name,))
    if not sp: return jsonify({"error": "Salesperson not found"}), 404
    sp_id = sp[0]
    cust_id = None
    if customer_name:
        c = one("SELECT id FROM customers WHERE name=%s", (customer_name,))
        if c: cust_id = c[0]

    # (optional) jamin invoice_no unik
    if one("SELECT 1 FROM sales_invoices WHERE invoice_no=%s", (invoice_no,)):
        return jsonify({"error":"invoice_no sudah dipakai"}), 409

    try:
        # Header
        g.con.execute("""
            INSERT INTO sales_invoices
            (salesperson_id, customer_id, invoice_no, invoice_date, due_date, status)
            VALUES (%s,%s,%s,%s,%s,'UNPAID')
        """, (sp_id, cust_id, invoice_no, invoice_date, due_date))
        sales_id = g.con.lastrowid

        total_hdr = Decimal('0')
        # Validasi & tulis item
        for it in items:
            product_id  = it.get('product_id') or it.get('id_barang')
            qty         = int(it.get('jmlpermintaan') or it.get('qty') or 0)
            unit_price  = parse_decimal(it.get('harga_satuan') or it.get('unit_price'))
            disc_pct    = parse_decimal(it.get('diskon') or it.get('discount_percent') or 0)
            total_amount= parse_decimal(it.get('harga_total') or it.get('total_amount'))

            # Hitung ulang jika perlu
            if total_amount == 0:
                gross = unit_price * qty
                total_amount = gross - (gross * disc_pct / Decimal('100'))

            # UoM-lite (kalau form tidak pakai, factor=1 & qty_base=qty)
            unit_label = it.get('unit_label')
            factor     = parse_decimal(it.get('uom_factor_to_base') or (1 if not unit_label else 1))
            qty_uom    = parse_decimal(it.get('qty_uom') or (qty if not unit_label else 0))
            qty_base   = parse_decimal(it.get('qty_base') or (qty_uom * factor if unit_label else qty))
            unit_price_uom  = parse_decimal(it.get('unit_price_uom') or (unit_price if not unit_label else 0))
            unit_price_base = parse_decimal(it.get('unit_price_base') or ((unit_price_uom/factor) if unit_label and factor else unit_price))

            # Cek stok tersedia (base unit)
            avail = one("""
                SELECT COALESCE(v.qty_on_hand,0) FROM v_product_stock v WHERE v.product_id=%s
            """, (product_id,))
            available = int(avail[0]) if avail else 0
            if int(qty_base) > available:
                return jsonify({"error": f"Stok {product_id} kurang. Tersedia {available}, diminta {int(qty_base)}"}), 400

            # Insert item
            g.con.execute("""
                INSERT INTO sales_items
                (sales_invoice_id, product_id, qty, unit_price, discount_percent, total_amount,
                 unit_label, uom_factor_to_base, qty_uom, qty_base, unit_price_uom, unit_price_base)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (sales_id, product_id, qty, unit_price, disc_pct, total_amount,
                  unit_label, factor, qty_uom, qty_base, unit_price_uom, unit_price_base))

            # Stock move (OUT)
            g.con.execute("""
                INSERT INTO stock_moves (product_id, ref_type, ref_id, qty_out, note)
                VALUES (%s,'SALE',%s,%s,%s)
            """, (product_id, sales_id, int(qty_base), f"Invoice {invoice_no}"))

            total_hdr += total_amount

        # (opsional) pembayaran awal
        paid_amount = parse_decimal(d.get('paid_amount') or 0)
        if paid_amount > 0:
            g.con.execute("""
                INSERT INTO payments (ref_type, ref_id, pay_date, method, amount, note)
                VALUES ('SALE', %s, %s, %s, %s, %s)
            """, (sales_id, invoice_date, d.get('method') or 'CASH', paid_amount, 'Pembayaran awal'))

        g.con.connection.commit()
        return jsonify({"msg":"SUKSES","id":sales_id,"total":str(total_hdr)})
    except Exception as e:
        g.con.connection.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/admin/pengeluaran/<int:id>', methods=['PUT'])
@jwt_required()
def pengeluaran_edit(id):
    d = request.get_json() or {}
    sp_id   = d.get('salesperson_id')
    cust_id = d.get('customer_id')
    invoice_no   = d.get('invoice_no')
    invoice_date = d.get('invoice_date')
    due_date     = d.get('due_date')
    items = d.get('items') or []
    if not (sp_id and invoice_no and invoice_date and items):
        return jsonify({"error":"salesperson_id, invoice_no, invoice_date, items wajib"}), 400
    try:
        g.con.execute("""
            UPDATE sales_invoices
            SET salesperson_id=%s, customer_id=%s, invoice_no=%s, invoice_date=%s, due_date=%s
            WHERE id=%s
        """, (sp_id, cust_id, invoice_no, invoice_date, due_date, id))

        g.con.execute("DELETE FROM stock_moves WHERE ref_type='SALE' AND ref_id=%s", (id,))
        g.con.execute("DELETE FROM sales_items WHERE sales_invoice_id=%s", (id,))

        total_hdr = Decimal('0')
        for it in items:
            product_id  = it.get('product_id')
            qty         = int(it.get('qty') or 0)
            unit_price  = parse_decimal(it.get('unit_price'))
            disc_pct    = parse_decimal(it.get('discount_percent') or 0)
            total_amount= parse_decimal(it.get('total_amount') or unit_price*qty - (unit_price*qty*disc_pct/Decimal('100')))

            unit_label = it.get('unit_label')
            factor     = parse_decimal(it.get('uom_factor_to_base') or (1 if not unit_label else 1))
            qty_uom    = parse_decimal(it.get('qty_uom') or (qty if not unit_label else 0))
            qty_base   = parse_decimal(it.get('qty_base') or (qty_uom * factor if unit_label else qty))

            avail = one("SELECT COALESCE(v.qty_on_hand,0) FROM v_product_stock v WHERE v.product_id=%s", (product_id,))
            available = int(avail[0]) if avail else 0
            if int(qty_base) > available:
                return jsonify({"error": f"Stok {product_id} kurang. Tersedia {available}, diminta {int(qty_base)}"}), 400

            g.con.execute("""
                INSERT INTO sales_items
                (sales_invoice_id, product_id, qty, unit_price, discount_percent, total_amount,
                 unit_label, uom_factor_to_base, qty_uom, qty_base)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (id, product_id, qty, unit_price, disc_pct, total_amount,
                  unit_label, factor, qty_uom, qty_base))

            g.con.execute("""
                INSERT INTO stock_moves (product_id, ref_type, ref_id, qty_out, note)
                VALUES (%s,'SALE',%s,%s,%s)
            """, (product_id, id, int(qty_base), f"Edit Invoice {invoice_no}"))

            total_hdr += total_amount

        g.con.connection.commit()
        return jsonify({"msg":"SUKSES","total":str(total_hdr)})
    except Exception as e:
        g.con.connection.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/admin/pengeluaran/<int:id>', methods=['DELETE'])
@jwt_required()
def pengeluaran_delete(id):
    try:
        g.con.execute("DELETE FROM stock_moves WHERE ref_type='SALE' AND ref_id=%s", (id,))
        g.con.execute("DELETE FROM sales_items WHERE sales_invoice_id=%s", (id,))
        g.con.execute("DELETE FROM sales_invoices WHERE id=%s", (id,))
        g.con.connection.commit()
        return jsonify({"msg":"BERHASIL DIHAPUS"})
    except Exception as e:
        g.con.connection.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/ceknofaktur/<path:tanggal>')
def ceknofaktur(tanggal):
    dt_tanggal = None
    print(tanggal)
    # Coba parse dengan 2 format umum
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            dt_tanggal = datetime.strptime(tanggal, fmt)
            break
        except ValueError:
            continue

    if not dt_tanggal:
        return {"error": "Format tanggal tidak dikenali (pakai YYYY-MM-DD atau DD/MM/YYYY)"}, 400

    # Hitung Jatuh Tempo tambah 1 bulan
    jth_tempo = (dt_tanggal + relativedelta(months=1)).strftime("%d/%m/%Y")

    # Query DB pakai format ISO standar
    tgl_str = dt_tanggal.strftime('%Y-%m-%d')
    g.con.execute("""
        SELECT RIGHT(nomerfaktur, 3) AS last_3
        FROM barang_keluar 
        WHERE tglfaktur = %s 
        ORDER BY nomerfaktur DESC LIMIT 1
    """, (tgl_str,))
    row = g.con.fetchone()
    last_3 = int(row[0]) if row and row[0] else 0

    # Gabungkan tanggal (yyyymmdd) + nomor urut
    nofaktur = f"{dt_tanggal.strftime('%Y%m%d')}{last_3 + 1:03d}"

    return jsonify({"jatuh_tempo": jth_tempo, "nofaktur": nofaktur})

def to_decimal(val):
    try:
        # pastikan val string dulu
        val = str(val).replace('.', '').replace(',', '.')
        return Decimal(val)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0.00')
def awal(pdf,width,height,barang_keluar,ada_diskon,ada_batch,ada_ed, cash_tempo):
        # Set margin awal
        x_margin = 0.75 * cm
        y = height - 1 * cm
        y2 = y
        # Header
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(x_margin, y - 0 * cm, "PT. BREGAS SAPORETE MEDIKALINDO")
        pdf.setFont("Helvetica", 8 )
        pdf.drawString(x_margin, y - 0.6 * cm, "Jl. Pati Perum Vero Permai No.39 Margadana, Kota Tegal")
        pdf.drawString(x_margin, y - 1 * cm, "Telp: 02833170260 | email : bregassaporetemedikalindo@gmail.com ") 
        pdf.drawString(x_margin, y - 1.4 * cm, "No Rek: 0101-01-003113-30-5 An PT Bregas Saporete Medikalindo(BRI)")
        pdf.drawString(x_margin, y - 1.8 * cm, "Izin CDAKB: PB-UMKU 912010045181900010001")
        y -= 2.4 * cm
        x_margin_9 = x_margin + 9.6 * cm   
        # Garis pemisah
        pdf.line(x_margin, y, width - x_margin, y)
        y -= 0.5 * cm

        # Info Faktur dan "Kepada Yth" (samping-sampingan)
        tanggal = barang_keluar[0].strftime('%d/%m/%Y')
        jth_tempo = barang_keluar[1].strftime('%d/%m/%Y')

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(x_margin_9, y2, "FAKTUR")
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(x_margin_9 + 4.2 * cm, y2, "Kepada Yth:")

        # Data Faktur di sebelah kiri
        pdf.setFont("Helvetica", 8)
        y2 -= 0.2 * cm
        pdf.drawString(x_margin_9, y2 - 0.4 * cm, "No Faktur")
        pdf.drawString(x_margin_9 + 1.9 * cm, y2 - 0.4 * cm, ": "+barang_keluar[2])
        pdf.drawString(x_margin_9, y2 - 0.8 * cm, "Inv Date")
        pdf.drawString(x_margin_9 + 1.9 * cm, y2 - 0.8 * cm, ": "+tanggal)
        pdf.drawString(x_margin_9, y2 - 1.2 * cm, "Sales")
        print(barang_keluar[3])
        pdf.drawString(x_margin_9 + 1.9 * cm, y2 - 1.2 * cm, ": "+barang_keluar[3])
        pdf.drawString(x_margin_9, y2 - 1.6 * cm, "Jatuh Tempo")
        pdf.drawString(x_margin_9 + 1.9 * cm, y2 - 1.6 * cm, ": "+jth_tempo)
        pdf.drawString(x_margin_9, y2 - 2 * cm, "Cash / Tempo")
        pdf.drawString(x_margin_9 + 1.9 * cm, y2 - 2 * cm, ": "+cash_tempo)
        
        # Data 'Kepada Yth' dan Alamat di sebelah kanan
        alamat = barang_keluar[5]
        alamat_lines = textwrap.wrap(alamat, width=50)
        pdf.drawString(x_margin_9 + 4.2 * cm, y2 - 0.4 * cm, barang_keluar[4])  # Nama outlet
        y3 = y2 
        for i, line in enumerate(alamat_lines[:11]):  # Limit to 6 lines
            y3 = y2 - (0.8 + i * 0.4) * cm
            pdf.drawString(x_margin_9 + 4.2 * cm, y2 - (0.8 + i * 0.4) * cm, line)  # Alamat outlet
        pdf.drawString(x_margin_9 + 4.2 * cm, y3 - 0.4 * cm, f"NPWP: {barang_keluar[6]}")  # Npwp outlet
        # Tabel barang
        pdf.setFont("Helvetica-Bold", 9)
        headers = ["No", "Nama Barang", "Qty"]
        header_positions = [x_margin, x_margin + 0.6 * cm, x_margin + 9 * cm]
        
        if ada_batch == True :
            header_positions[2] =  x_margin + 7 * cm
            headers.append("Batch")
            if ada_ed == True :
                header_positions[2] =  x_margin + 6 * cm
                header_positions.append(x_margin + 7.5 * cm)
            else:
                header_positions.append(x_margin + 8 * cm)  
        if ada_ed == True:
            headers.append("ED")
            if ada_batch == True :
                header_positions.append(x_margin + 9.5 * cm)  
            else:
                header_positions[2] =  x_margin + 7 * cm
                header_positions.append(x_margin + 8 * cm)  
        headers.append("Harga")
        header_positions.append(x_margin + 11 * cm) 
        if ada_diskon == True:
            headers.append("Diskon")
            header_positions.append(x_margin + 14.5 * cm)
        headers.append("Total")
        header_positions.append(x_margin + 16 * cm)
        for pos, header in zip(header_positions, headers):
            pdf.drawString(pos, y, header)
        y -= 0.3 * cm
        pdf.line(x_margin, y, width - x_margin, y)
        y -= 0.5 * cm

        pdf.setFont("Helvetica", 9)
        return y, x_margin,width, header_positions
    
    # footer total
def hitung_total(pdf, y, x_margin, width, jumlah_total, pajak, page_num=1, total_pages=1, ada_diskon=False ):
    if page_num == total_pages:
        pdf.line(x_margin, y, width - x_margin, y)
        y -= 0.4 * cm
        if pajak == "pajak":
            pdf.setFont("Helvetica", 8)
            if ada_diskon == True:
                pdf.drawString(x_margin + 15.975 * cm, y, f"Rp ")
                pdf.drawString(x_margin + 15.975 * cm, y - 0.4 * cm, f"Rp ")  
                pdf.drawString(x_margin + 15.975 * cm, y - 0.8 * cm, f"Rp ")   
            else:
                pdf.drawString(x_margin + 16 * cm, y, f"Rp ")
                pdf.drawString(x_margin + 16 * cm, y - 0.4 * cm, f"Rp ")  
                pdf.drawString(x_margin + 16 * cm, y - 0.8 * cm, f"Rp ")     
            pdf.drawString(x_margin + 14 * cm, y, f"Total: ")
            pdf.drawRightString(x_margin + 19.6 * cm, y, f"{format_rupiah(jumlah_total)}")
            pajak_val = Decimal('0.11') * jumlah_total
            pajak_val = pajak_val.quantize(Decimal('1'), rounding=decimal.ROUND_HALF_UP) # pembulatan
            pdf.drawString(x_margin + 14 * cm, y - 0.4 * cm, f"PPN 11%: ")  
            pdf.drawRightString(x_margin + 19.6 * cm, y - 0.4 * cm, f"{format_rupiah(pajak_val)}")
            grand_total = jumlah_total + pajak_val
            terbilang_rupiah = num2words(int(grand_total), lang='id')
            terbilang_rupiah = terbilang_rupiah.title() 
            terbilang_rupiah += " Rupiah"
            pdf.setFont("Helvetica-Bold", 8)
            y3 = y
            if len(terbilang_rupiah) >= 90:
                terbilang_rupiah_lines = textwrap.wrap(terbilang_rupiah, width=75)
                for i, line in enumerate(terbilang_rupiah_lines[:3]):
                    y3 -=  0.4 * cm
                    pdf.drawString(x_margin ,y - (i * 0.4) * cm, line)
            else:
                pdf.drawString(x_margin, y , terbilang_rupiah)
            pdf.drawString(x_margin + 14 * cm, y - 0.8 * cm, f"Grand Total: ")
            
            pdf.drawRightString(x_margin + 19.6 * cm, y - 0.8 * cm, f"{format_rupiah(grand_total)}")

            # Footer
            pdf.drawString(x_margin, y3 - 0.5 * cm, "Penerima")
            pdf.drawString(x_margin + 4.5 * cm, y3 - 0.5 * cm, "Gudang")
            pdf.drawString(x_margin + 8.5 * cm, y3 - 0.5 * cm, "Expedisi")
        else:
            pdf.setFont("Helvetica-Bold", 10)
            terbilang_rupiah = num2words(jumlah_total, lang='id')
            terbilang_rupiah = terbilang_rupiah.title() 
            terbilang_rupiah += " Rupiah"
            y3 = y
            if len(terbilang_rupiah) >= 90:
                terbilang_rupiah_lines = textwrap.wrap(terbilang_rupiah, width=75)
                for i, line in enumerate(terbilang_rupiah_lines[:3]):
                    y3 -=  0.4 * cm
                    pdf.drawString(x_margin ,y - (i * 0.4) * cm, line)
            else:
                pdf.drawString(x_margin, y , terbilang_rupiah)
            if ada_diskon == True:
                pdf.drawString(x_margin + 14.9 * cm, y, f"Total: Rp ")
            else:
                pdf.drawString(x_margin + 14.92 * cm, y, f"Total: Rp ")
            pdf.drawRightString(x_margin + 19.6 * cm, y, f"{format_rupiah(jumlah_total)}")
            y3 -= 0.5 * cm

            # Footer
            pdf.drawString(x_margin, y3, "Penerima")
            pdf.drawString(x_margin + 4.5 * cm, y3, "Gudang")
            pdf.drawString(x_margin + 8.5 * cm, y3, "Expedisi")
    else:
        pdf.line(x_margin, y, width - x_margin, y)
        y -= 0.4 * cm
        if pajak == "pajak":
            pdf.setFont("Helvetica-Bold", 8)
            y3 = y

            # Footer
            pdf.drawString(x_margin, y3 - 0.5 * cm, "Penerima")
            pdf.drawString(x_margin + 4.5 * cm, y3 - 0.5 * cm, "Gudang")
            pdf.drawString(x_margin + 8.5 * cm, y3 - 0.5 * cm, "Expedisi")
        else:
            pdf.setFont("Helvetica-Bold", 10)
            y3 = y
            y3 -= 0.5 * cm

            # Footer
            pdf.drawString(x_margin, y3, "Penerima")
            pdf.drawString(x_margin + 4.5 * cm, y3, "Gudang")
            pdf.drawString(x_margin + 8.5 * cm, y3, "Expedisi")

def export_pdf(buffer, pajak):
    id = request.args.get("id")
    g.con.execute("""SELECT invoice_date, due_date, invoice_no, note
    FROM purchases WHERE purcases.id = %s """, (id,))
    barang_keluar = g.con.fetchone()
    data_db = g.con.fetchall()
    data = []
    jumlah_total = 0
    ada_diskon = False
    ada_batch = False
    ada_ed = False
    for i in range(len(data_db)):
        item = {"No":i+1, "Nama Barang":data_db[i][0],"Quantity":str(data_db[i][1])+" "+data_db[i][2],
                "Harga Satuan": format_currency(data_db[i][3]),"Total":format_currency(data_db[i][4])}
        data.append(item)
        jumlah_total += data_db[i][4]
        if data_db[i][5] and data_db[i][5] != '' and data_db[i][5] != 0:
            ada_diskon = True
        if data_db[i][6] and data_db[i][6] != '' and data_db[i][6] != 0:
            ada_batch = True
        if data_db[i][7] and data_db[i][7] != '' and data_db[i][7] != 0:
            ada_ed = True

    # Pagination logic
    max_rows_per_page = 12
    total_rows = max(len(data_db), max_rows_per_page)
    total_pages = (total_rows + max_rows_per_page - 1) // max_rows_per_page

    pdf = canvas.Canvas(buffer, pagesize=(21.6 * cm, 14.5 * cm))
    width, height = (21.6 * cm, 14.5 * cm)
    cash_tempo = barang_keluar[7]
    y, x_margin, width, header_positions = awal(
                pdf, width, height, barang_keluar, ada_diskon, ada_batch, ada_ed, cash_tempo
            )

    jumlah_total = 0
    row_idx = 0
    page_num = 1

    for i in range(total_rows):
        if (i > 0 and i % max_rows_per_page == 0):
            hitung_total(pdf, y, x_margin, width, jumlah_total, pajak, page_num, total_pages,ada_diskon)
            pdf.showPage()
            page_num += 1
            y, x_margin, width, header_positions = awal(
                pdf, width, height, barang_keluar, ada_diskon, ada_batch, ada_ed, cash_tempo
            )

        if i < len(data_db):
            nama_barang = data_db[i][0]
            jml = data_db[i][1]
            satuan = data_db[i][2]
            harga_satuan = data_db[i][3]
            total = data_db[i][4]
            diskon = data_db[i][5] if ada_diskon else None
            batch = data_db[i][6] if ada_batch else None
            ed = data_db[i][7] if ada_ed else None

            jumlah_total += total
            qty = f"{jml} {satuan}"

            values = [str(i + 1), nama_barang, qty]

            if ada_batch:
                values.append(str(batch) if batch else "-")
            if ada_ed:
                values.append(str(ed) if ed else "-")

            values.append(f"satuan {format_rupiah(harga_satuan)}")

            if ada_diskon:
                values.append(f"{diskon}%")
            
            values.append(f"total {format_rupiah(total)}")

        else:
            # Isi kosong untuk baris tidak terpakai
            values = [str(i + 1), "-", "-"]
            if ada_batch:
                values.append("-")
            if ada_ed:
                values.append("-")
            values.append("-")  # harga
            if ada_diskon:
                values.append("-")
            values.append("-")  # total
        # Gambar data ke PDF
        for pos, value in zip(header_positions, values):
            if value.startswith("satuan"):
                pdf.drawString(pos, y, "Rp ")
                pdf.drawRightString(pos + (3.2 * cm), y, value[7:].strip())
            elif value.startswith("total"):
                pdf.drawString(pos, y, "Rp ")
                pdf.drawRightString(x_margin + 19.6 * cm, y, value[6:].strip())
            else:
                pdf.drawString(pos, y, value)
        y -= 0.5 * cm

    print(jumlah_total)
    # Total
    hitung_total(pdf, y, x_margin,width, jumlah_total, pajak, page_num, total_pages, ada_diskon)
    #return Response(generate(), mimetype='application/pdf')
    pdf.save()

@app.route('/admin/pengeluaran/print', methods=['GET'])
def print_pdf():
    buffer = io.BytesIO()
    export_pdf(buffer,"")
    buffer.seek(0)
    print(f"PDF size: {len(buffer.getvalue())} bytes")  # Debug ukuran file
    def generate():
        yield buffer.read()
    return send_file(buffer,as_attachment=True,download_name="faktur.pdf",mimetype='application/pdf')

@app.route('/admin/pengeluaran/print_pajak', methods=['GET'])
def print_pdf_pajak():
    buffer = io.BytesIO()
    export_pdf(buffer,"pajak")
    buffer.seek(0)
    print(f"PDF size: {len(buffer.getvalue())} bytes")  # Debug ukuran file
    def generate():
        yield buffer.read()
    return send_file(buffer,as_attachment=True,download_name="faktur+pajak.pdf",mimetype='application/pdf')
@app.route('/admin/pengeluaran/hapus/id', methods=['DELETE'])
@jwt_required()
def hapus_pengeluaran():
    try:
        id = request.json.get('id')
        print(id)
        g.con.execute("SELECT id_barang, jmlpermintaan, id FROM detail_barang_keluar WHERE id_barang_keluar = %s", 
        (int(id),))
        data = g.con.fetchall()
        print(data)
        if not data:
            return jsonify({"error": "Data tidak ditemukan"}), 404

        for i in data:
            # Kurangi jumlah di barang_gudang
            g.con.execute("SELECT sisa_gudang FROM barang_gudang WHERE id_barang = %s", (i[0],))
            gudang = g.con.fetchone()
            g.con.execute("SELECT stoklimit FROM barang WHERE id = %s", (i[0],))
            stoklimit = g.con.fetchone()
            if gudang:
                new_jumlah = max(0, gudang[0] + i[1])
                if int(new_jumlah) <= stoklimit[0]:
                    ket = "Stok Tidak Aman"
                else :
                    ket = "Stok Aman"
                g.con.execute("UPDATE barang_gudang SET sisa_gudang = %s, keterangan = %s WHERE id_barang = %s"
                              , (new_jumlah, ket, i[0]))
            # Hapus dari detail_barang_masuk
            g.con.execute("DELETE FROM detail_barang_keluar WHERE id = %s", (i[2],))
        # Hapus dari barang_masuk
        g.con.execute("DELETE FROM barang_keluar WHERE id = %s", (int(id),))
        g.con.connection.commit()
        return jsonify({"msg": "SUKSES"})
    except Exception as e:
        print(str(e))
        return jsonify({"error": str(e)})
@app.route('/admin/keuangan')
def adminkeuangan():
    # --- Params ---
    tahun        = request.args.get('tahun', type=int)
    bulan        = request.args.get('bulan', type=int)
    tanggal      = request.args.get('tanggal', type=int)
    nama_sales_q = request.args.get('nama_sales', '', type=str)     # di UI boleh isi nama
    nama_cust_q  = request.args.get('nama_outlet', '', type=str)    # di UI boleh isi nama
    page         = request.args.get('page', default=1, type=int)
    per_page     = request.args.get('per_page', default=10, type=int)

    # --- Build filter WHERE ---
    filters, params, count_params = [], [], []
    clause, prms = build_date_range(
        year=tahun, month=bulan, day=tanggal,
        alias="si", col="invoice_date"
    )
    if clause:
        filters.append(clause); params += prms; count_params += prms
    if nama_sales_q:
        filters.append("s.name = %s"); params.append(nama_sales_q); count_params.append(nama_sales_q)
    if nama_cust_q:
        filters.append("c.name = %s"); params.append(nama_cust_q); count_params.append(nama_cust_q)

    where_clause = (" WHERE " + " AND ".join(filters)) if filters else ""

    # --- Hitung total records (distinct invoice) ---
    count_sql = f"""
        SELECT COUNT(*) FROM (
          SELECT si.id
          FROM sales_invoices si
          JOIN salespersons s ON s.id = si.salesperson_id
          LEFT JOIN customers  c ON c.id = si.customer_id
          {where_clause}
          GROUP BY si.id
        ) x
    """
    g.con.execute(count_sql, tuple(count_params))
    total_records = g.con.fetchone()[0] or 0

    # --- Ambil list invoice + agregat (total invoice, total bayar, outstanding) ---
    offset = (page - 1) * per_page
    invoices = fetch(f"""
        SELECT
          si.id,
          si.invoice_date   AS tglfaktur,
          si.invoice_no     AS nomerfaktur,
          si.due_date       AS jatuhtempo,
          si.status,
          s.name            AS nama_sales,
          COALESCE(c.name,'-')    AS nama_outlet,
          COALESCE(c.address,'')  AS alamat_outlet,
          COALESCE(c.npwp,'')     AS npwp,
          COALESCE(si_amt.total_invoice, 0) AS total_invoice,
          COALESCE(pay_amt.total_payment, 0) AS total_payment,
          COALESCE(si_amt.total_invoice, 0) - COALESCE(pay_amt.total_payment, 0) AS outstanding,
          -- metode bayar terakhir (opsional, meniru cashtempo lama)
          (SELECT p2.method FROM payments p2
             WHERE p2.ref_type='SALE' AND p2.ref_id=si.id
             ORDER BY p2.pay_date DESC, p2.id DESC LIMIT 1) AS last_pay_method,
          (SELECT p3.pay_date FROM payments p3
             WHERE p3.ref_type='SALE' AND p3.ref_id=si.id
             ORDER BY p3.pay_date DESC, p3.id DESC LIMIT 1) AS last_pay_date
        FROM sales_invoices si
        JOIN salespersons s ON s.id = si.salesperson_id
        LEFT JOIN customers  c ON c.id = si.customer_id
        LEFT JOIN (
            SELECT sales_invoice_id, SUM(total_amount) AS total_invoice
            FROM sales_items
            GROUP BY sales_invoice_id
        ) si_amt ON si_amt.sales_invoice_id = si.id
        LEFT JOIN (
            SELECT ref_id, SUM(amount) AS total_payment
            FROM payments
            WHERE ref_type='SALE'
            GROUP BY ref_id
        ) pay_amt ON pay_amt.ref_id = si.id
        {where_clause}
        ORDER BY si.id DESC
        LIMIT %s OFFSET %s
    """, tuple(params + [per_page, offset]))

    # --- Ambil detail item untuk invoice di halaman ini (hindari N+1) ---
    detail_map = {}
    if invoices:
        ids = tuple(r['id'] for r in invoices)
        # MySQL butuh IN (%s) dengan tuple; kalau single, pastikan tuple satu elemen (id,)
        detail_rows = fetch("""
            SELECT
              it.sales_invoice_id,
              p.code AS kode_barang,
              p.name AS nama_barang,
              p.unit AS unit,
              it.qty,
              it.unit_price,
              it.discount_percent,
              it.total_amount
            FROM sales_items it
            JOIN products p ON p.id = it.product_id
            WHERE it.sales_invoice_id IN %s
            ORDER BY it.id
        """, (ids,))
        for d in detail_rows:
            detail_map.setdefault(d['sales_invoice_id'], []).append(d)

    # --- Susun payload untuk template ---
    info_list = []
    for inv in invoices:
        info_list.append({
            "id_barang_keluar": inv["id"],           # kompatibel dgn template lama
            "tglfaktur": inv["tglfaktur"],
            "nomerfaktur": inv["nomerfaktur"],
            "jatuhtempo": inv["jatuhtempo"],
            "nama_sales": inv["nama_sales"],
            "nama_outlet": inv["nama_outlet"],
            "alamat_outlet": inv["alamat_outlet"],
            "npwp": inv["npwp"],
            "status": inv["status"],
            "last_pay_method": inv["last_pay_method"] or "-",
            "last_pay_date": inv["last_pay_date"] or None,
            "total_invoice": inv["total_invoice"],
            "total_payment": inv["total_payment"],
            "outstanding": inv["outstanding"],
            "detail_items": detail_map.get(inv["id"], [])
        })

    # --- Lookup untuk filter dropdown ---
    data_sales = fetch("SELECT id, name FROM salespersons WHERE is_active=1 ORDER BY name")
    data_outlet = fetch("SELECT id, name, address FROM customers ORDER BY name")

    # Data pendukung lain (opsional; sesuaikan template kalau tidak dipakai)
    # tanggal list (distinct invoice_date)
    tanggal_pengeluaran = fetch("""
        SELECT DISTINCT invoice_date AS tglfaktur
        FROM sales_invoices
        ORDER BY invoice_date DESC
    """)

    # No faktur & jatuh tempo default (untuk modal tambah cepat)
    today = time_zone_wib().date()
    nofaktur  = next_invoice_no_for_date(today)
    jth_tempo = (today + timedelta(days=30)).strftime("%Y-%m-%d")

    # --- Pagination info ---
    total_pages = (total_records + per_page - 1) // per_page
    has_next = page < total_pages
    has_prev = page > 1
    page_range = range(max(1, page - 3), min(total_pages + 1, page + 3))

    return render_pjax(
        "admin/keuangan.html",
        info_list=info_list,
        # filters & dropdowns
        tahun=fetch("SELECT DISTINCT YEAR(invoice_date) AS y FROM sales_invoices ORDER BY y DESC"),
        data_sales=data_sales,
        data_outlet=data_outlet,
        nama_outlet=fetch("SELECT DISTINCT name FROM customers ORDER BY name"),
        nama_sales=fetch("SELECT DISTINCT name FROM salespersons WHERE is_active=1 ORDER BY name"),
        tanggal_pengeluaran=tanggal_pengeluaran,
        # quick-add helpers
        tanggal=today.strftime("%Y-%m-%d"),
        nofaktur=nofaktur,
        jth_tempo=jth_tempo,
        # pager
        page=page, per_page=per_page,
        total_pages=total_pages, total_records=total_records,
        has_next=has_next, has_prev=has_prev, page_range=page_range
    ) 
@app.route('/admin/keuangan/edit/id', methods=['PUT'])
@jwt_required()
def keuangan_edit():
    d = request.get_json() or {}
    # dukung nama field lama & baru
    sales_id   = d.get('id') or d.get('id_sales') or d.get('id_barang_keluar')
    invoice_no = d.get('invoice_no') or d.get('nofaktur') or d.get('nomerfaktur')
    pay_date   = d.get('pay_date') or d.get('tanggal_pembayaran')        # 'YYYY-MM-DD'
    method     = d.get('method') or d.get('metode_pembayaran') or d.get('cashtempo')  # 'CASH'/'TRANSFER'/dst
    amount     = d.get('amount')  # nominal pembayaran (Decimal/str/number)
    note       = d.get('note') or d.get('keterangan_pembayaran') or ''
    # map Lunas/Tidak Lunas lama -> status baru (opsional)
    lunas_tidak = (d.get('lunas_tidak') or '').strip().lower()
    explicit_status = d.get('status')
    if not explicit_status and lunas_tidak:
        explicit_status = 'PAID' if 'lunas' in lunas_tidak else 'UNPAID'

    if not sales_id:
        return jsonify({"error": "id (sales invoice) wajib"}), 400

    try:
        g.con.execute("START TRANSACTION")

        # pastikan invoice ada
        g.con.execute("SELECT id FROM sales_invoices WHERE id=%s", (sales_id,))
        if not g.con.fetchone():
            g.con.execute("ROLLBACK")
            return jsonify({"error": "Sales invoice tidak ditemukan"}), 404

        # update nomor faktur jika ada
        if invoice_no:
            g.con.execute("UPDATE sales_invoices SET invoice_no=%s WHERE id=%s",
                          (invoice_no, sales_id))

        # tambah pembayaran jika diberikan tanggal & amount
        if pay_date and amount not in (None, '',):
            try:
                amt = Decimal(str(amount))
            except (InvalidOperation, ValueError, TypeError):
                g.con.execute("ROLLBACK")
                return jsonify({"error": "amount tidak valid"}), 400

            g.con.execute("""
                INSERT INTO payments (ref_type, ref_id, pay_date, method, amount, note)
                VALUES ('SALE', %s, %s, %s, %s, %s)
            """, (sales_id, pay_date, (method or 'CASH'), amt, note))

        # set status eksplisit jika dikirim
        if explicit_status:
            g.con.execute("UPDATE sales_invoices SET status=%s WHERE id=%s",
                          (explicit_status.upper(), sales_id))

        # jika tidak ada status eksplisit, hitung otomatis dari total vs pembayaran
        if not explicit_status:
            # total invoice
            g.con.execute("""
                SELECT COALESCE(SUM(total_amount),0) FROM sales_items
                WHERE sales_invoice_id=%s
            """, (sales_id,))
            total_inv = g.con.fetchone()[0] or 0

            # total pembayaran
            g.con.execute("""
                SELECT COALESCE(SUM(amount),0) FROM payments
                WHERE ref_type='SALE' AND ref_id=%s
            """, (sales_id,))
            total_pay = g.con.fetchone()[0] or 0

            new_status = 'UNPAID'
            if total_pay <= 0:
                new_status = 'UNPAID'
            elif total_pay < total_inv:
                new_status = 'PARTIAL'
            else:
                new_status = 'PAID'

            g.con.execute("UPDATE sales_invoices SET status=%s WHERE id=%s",
                          (new_status, sales_id))

        g.con.execute("COMMIT")
        return jsonify({"msg": "SUKSES"})
    except Exception as e:
        g.con.execute("ROLLBACK")
        return jsonify({"error": str(e)}), 500
@app.route('/api/payments/sale/<int:ref_id>')
def get_payments(ref_id):
    rows = fetch("""
        SELECT id, pay_date, method, amount, note
        FROM payments
        WHERE ref_type='SALE' AND ref_id=%s
        ORDER BY pay_date
    """, (ref_id,))
    return jsonify(rows)

@app.route('/api/payments/add', methods=['POST'])
@jwt_required()
def add_payment():
    d = request.json
    g.con.execute("""
        INSERT INTO payments (ref_type, ref_id, pay_date, method, amount, note)
        VALUES ('SALE', %s, %s, %s, %s, %s)
    """, (d['ref_id'], d['pay_date'], d['method'], d['amount'], d.get('note','')))
    g.con.connection.commit()
    return jsonify({'msg':'ok'})

@app.route('/api/payments/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_payment(id):
    g.con.execute("DELETE FROM payments WHERE id=%s", (id,))
    g.con.connection.commit()
    
@app.route('/admin/administrasi')
def adminadministrasi():
    tahun   = request.args.get('tahun', type=int)
    bulan   = request.args.get('bulan', type=int)
    tanggal = request.args.get('tanggal', type=int)
    supplier_name = request.args.get('nama_principle', type=str)
    page     = request.args.get('page', default=1, type=int)
    per_page = request.args.get('per_page', default=10, type=int)

    filters, params, count_params = [], [], []

    clause, prms = build_date_range(
        year=tahun, month=bulan, day=tanggal,
        alias="p", col="invoice_date"
    )
    if clause:
        filters.append(clause); params += prms; count_params += prms
    if supplier_name:
        filters.append("s.name = %s"); params.append(supplier_name); count_params.append(supplier_name)

    where_clause = (" WHERE " + " AND ".join(filters)) if filters else ""

    # Count distinct purchases
    cnt = one(f"""
        SELECT COUNT(*) FROM (
          SELECT p.id
          FROM purchases p
          JOIN suppliers s ON s.id = p.supplier_id
          {where_clause}
          GROUP BY p.id
        ) x
    """, tuple(count_params))
    total_records = cnt[0] if cnt else 0

    offset = (page - 1) * per_page

    # Header + agregat total & payment
    rows = fetch(f"""
        SELECT
          p.id,
          p.invoice_date AS tglfaktur,
          p.invoice_no   AS nofaktur,
          p.status,
          s.name   AS nama_supplier,
          s.address AS alamat,
          s.phone   AS tlp,
          COALESCE(pi_sum.total_belanja, 0)  AS performa_belanja,
          COALESCE(pay_sum.total_bayar, 0)   AS total_bayar,
          (COALESCE(pi_sum.total_belanja, 0) - COALESCE(pay_sum.total_bayar, 0)) AS outstanding,
          -- pembayaran terakhir (optional, untuk tampilan)
          (SELECT p2.pay_date FROM payments p2
             WHERE p2.ref_type='PURCHASE' AND p2.ref_id=p.id
             ORDER BY p2.pay_date DESC, p2.id DESC LIMIT 1) AS tanggal_pembayaran_terakhir,
          (SELECT p3.method FROM payments p3
             WHERE p3.ref_type='PURCHASE' AND p3.ref_id=p.id
             ORDER BY p3.pay_date DESC, p3.id DESC LIMIT 1) AS metode_terakhir,
          (SELECT p4.note FROM payments p4
             WHERE p4.ref_type='PURCHASE' AND p4.ref_id=p.id
             ORDER BY p4.pay_date DESC, p4.id DESC LIMIT 1) AS keterangan_terakhir
        FROM purchases p
        JOIN suppliers s ON s.id = p.supplier_id
        LEFT JOIN (
          SELECT purchase_id, SUM(total_amount) AS total_belanja
          FROM purchase_items
          GROUP BY purchase_id
        ) pi_sum ON pi_sum.purchase_id = p.id
        LEFT JOIN (
          SELECT ref_id, SUM(amount) AS total_bayar
          FROM payments
          WHERE ref_type='PURCHASE'
          GROUP BY ref_id
        ) pay_sum ON pay_sum.ref_id = p.id
        {where_clause}
        ORDER BY p.id DESC
        LIMIT %s OFFSET %s
    """, tuple(params + [per_page, offset]))

    # Detail item untuk invoice di halaman ini
    detail_map = {}
    if rows:
        ids = tuple(r['id'] for r in rows)
        detail = fetch("""
            SELECT
              pi.purchase_id,
              pr.code, pr.name,
              pi.qty, pi.unit_price, pi.total_amount,
              pi.unit_label, pi.qty_uom, pi.uom_factor_to_base, pi.qty_base
            FROM purchase_items pi
            JOIN products pr ON pr.id = pi.product_id
            WHERE pi.purchase_id IN %s
            ORDER BY pi.id
        """, (ids,))
        for d in detail:
            detail_map.setdefault(d['purchase_id'], []).append(d)

    info_list = []
    for r in rows:
        info_list.append({
            "id": r["id"],
            "tglfaktur": r["tglfaktur"],
            "nofaktur": r["nofaktur"],
            "status": r["status"],
            "nama_supplier": r["nama_supplier"],
            "alamat": r["alamat"],
            "tlp": r["tlp"],
            "performa_belanja": r["performa_belanja"],
            "total_bayar": r["total_bayar"],
            "outstanding": r["outstanding"],
            "tanggal_pembayaran_terakhir": r["tanggal_pembayaran_terakhir"],
            "metode_terakhir": r["metode_terakhir"],
            "keterangan_terakhir": r["keterangan_terakhir"],
            "detail_items": detail_map.get(r["id"], [])
        })

    # Dropdown/filter data
    data_pabrik = fetch("SELECT id, name AS nama_supplier, address AS alamat, phone AS tlp FROM suppliers ORDER BY name")
    # daftar tahun untuk filter (distinct)
    tahun_list = fetch("SELECT DISTINCT YEAR(invoice_date) AS tahun FROM purchases ORDER BY tahun DESC")

    # pagination meta
    total_pages = (total_records + per_page - 1) // per_page
    has_next = page < total_pages
    has_prev = page > 1
    page_range = range(max(1, page - 3), min(total_pages + 1, page + 3))

    return render_pjax(
        "admin/administrasi.html",
        data_pabrik=data_pabrik,
        info_list=info_list,
        list_tahun=tahun_list,
        page=page, per_page=per_page,
        total_pages=total_pages, total_records=total_records,
        has_next=has_next, has_prev=has_prev, page_range=page_range
    )
@app.route('/admin/administrasi/edit/id', methods=['PUT'])
@jwt_required()
def edit_administrasi():
    d = request.get_json() or {}
    purchase_id = d.get('id_purchase') or d.get('id') or d.get('id_barang_masuk')
    if not purchase_id:
        return jsonify({"error":"id_purchase wajib"}), 400

    status = d.get('status')  # optional: 'POSTED'/'PAID'/'PARTIAL' dll.
    pay_date = d.get('pay_date') or d.get('tanggal_pembayaran')  # YYYY-MM-DD (opsional)
    method   = d.get('method') or d.get('metode_pembayaran')     # 'TRANSFER'/'CASH'/dst (opsional)
    amount   = d.get('amount')
    note     = d.get('note') or d.get('keterangan_pembayaran')

    try:
        g.con.execute("START TRANSACTION")

        # update status jika ada
        if status:
            g.con.execute("UPDATE purchases SET status=%s WHERE id=%s", (status, purchase_id,))

        # tambah pembayaran jika ada fieldnya
        if pay_date and amount is not None:
            g.con.execute("""
                INSERT INTO payments (ref_type, ref_id, pay_date, method, amount, note)
                VALUES ('PURCHASE', %s, %s, %s, %s, %s)
            """, (purchase_id, pay_date, method or 'TRANSFER', D(amount), note or ''))

        g.con.connection.commit()
        return jsonify({"msg":"SUKSES"})
    except Exception as e:
        g.con.connection.rollback()
        return jsonify({"error": str(e)}), 500
@app.route('/admin/administrasi/hapus/id', methods=['DELETE'])
@jwt_required()
def hapus_administrasi():
    d = request.get_json() or {}
    purchase_id = d.get('id')
    if not purchase_id:
        return jsonify({"error":"id wajib"}), 400
    try:
        g.con.execute("START TRANSACTION")
        # hapus pergerakan stok yang terkait pembelian ini
        g.con.execute("DELETE FROM stock_moves WHERE ref_type='PURCHASE' AND ref_id=%s", (purchase_id,))
        # hapus detail
        g.con.execute("DELETE FROM purchase_items WHERE purchase_id=%s", (purchase_id,))
        # (opsional) hapus pembayaran yang terkait invoice ini
        g.con.execute("DELETE FROM payments WHERE ref_type='PURCHASE' AND ref_id=%s", (purchase_id,))
        # hapus header
        g.con.execute("DELETE FROM purchases WHERE id=%s", (purchase_id,))
        g.con.connection.commit()
        return jsonify({"msg":"SUKSES"})
    except Exception as e:
        g.con.connection.rollback()
        return jsonify({"error": str(e)}), 500
def query_performa(tahun: int | None, bulan: int | None,
                   sales_name: str | None, lunas_tidak: str | None):
    where, params = [], []

    # Filter tanggal berdasarkan header sales_invoices.invoice_date
    date_sql, date_params = build_date_range(
        year=tahun, month=bulan, day=None, alias="si", col="invoice_date"
    )
    if date_sql:
        where.append(date_sql)
        params.extend(date_params)

    # Filter nama sales (case-insensitive)
    if sales_name:
        where.append("UPPER(TRIM(s.name)) = UPPER(TRIM(%s))")
        params.append(sales_name)

    # Normalisasi status
    st = (lunas_tidak or '').strip().casefold()

    # Subquery per invoice: total_invoice & total_payment
    # → aman untuk ONLY_FULL_GROUP_BY
    sql = f"""
    WITH inv AS (
      SELECT
        si.id                AS id_faktur,
        s.id                 AS id_sales,
        s.name               AS nama_sales,
        COALESCE(c.name,'-') AS nama_outlet,
        YEAR(si.invoice_date)  AS tahun,
        MONTH(si.invoice_date) AS bulan,
        COALESCE(it_sum.total_invoice, 0) AS total_inv,
        COALESCE(pay_sum.total_payment, 0) AS total_pay
      FROM sales_invoices si
      JOIN salespersons s ON s.id = si.salesperson_id
      LEFT JOIN customers  c ON c.id = si.customer_id
      LEFT JOIN (
        SELECT sales_invoice_id, SUM(total_amount) AS total_invoice
        FROM sales_items
        GROUP BY sales_invoice_id
      ) it_sum ON it_sum.sales_invoice_id = si.id
      LEFT JOIN (
        SELECT ref_id, SUM(amount) AS total_payment
        FROM payments
        WHERE ref_type = 'SALE'
        GROUP BY ref_id
      ) pay_sum ON pay_sum.ref_id = si.id
      {"WHERE " + " AND ".join(where) if where else ""}
    )
    SELECT
      x.id_sales,
      x.nama_sales,
      x.nama_outlet,
      x.tahun,
      x.bulan,
      SUM(x.nilai) AS nilai
    FROM (
      SELECT
        id_faktur,
        id_sales,
        nama_sales,
        nama_outlet,
        tahun,
        bulan,
        CASE
          { "WHEN total_pay >= total_inv THEN total_inv" if st == "lunas" else
            "WHEN total_pay  < total_inv THEN total_inv" if st in ("tidak lunas","tdk lunas","belum lunas") else
            "WHEN 1=1 THEN total_inv" }
        END AS nilai
      FROM inv
    ) x
    WHERE x.nilai IS NOT NULL
    GROUP BY x.id_sales, x.nama_sales, x.nama_outlet, x.tahun, x.bulan
    ORDER BY x.tahun DESC, x.bulan ASC, x.nama_outlet, x.nama_sales;
    """
    return fetch(sql, tuple(params))

@app.route('/admin/performa_sales')
def adminperformasales():
    tahun = request.args.get('tahun', type=int)
    bulan = request.args.get('bulan', type=int)
    lunas_tidak = request.args.get('lunas_tidak', type=str)
    selected_sales = request.args.get('sales')

    # Ambil data agregat
    rows = query_performa(tahun, bulan, selected_sales, lunas_tidak)

    # Daftar tahun (untuk filter)
    list_tahun = fetch("""
        SELECT DISTINCT YEAR(invoice_date) AS tahun
        FROM sales_invoices
        ORDER BY tahun DESC
    """)
    thn = fetch("""
        SELECT DISTINCT YEAR(invoice_date) AS tahun
        FROM sales_invoices
        {where}
        ORDER BY tahun DESC
    """.format(where="WHERE YEAR(invoice_date)=%s" if tahun else ""), ([tahun] if tahun else []))

    # Dropdown nama sales
    nama_sales = fetch("SELECT DISTINCT name AS nama_sales FROM salespersons WHERE is_active=1 ORDER BY name")

    # Pivot: per (tahun, id_sales, outlet, sales)
    months = [bulan] if bulan else list(range(1, 13))
    idx = {}  # key: (tahun, id_sales, outlet, sales)

    for r in rows:
        key = (r['tahun'], r['id_sales'], r['nama_outlet'], r['nama_sales'])
        if key not in idx:
            idx[key] = {
                "tahun": r['tahun'],
                "id_sales": r['id_sales'],
                "nama_outlet": r['nama_outlet'],
                "nama_sales": r['nama_sales'],
                **{f"M{i}": 0 for i in months},
                "total_sales": 0
            }
        val = int(r['nilai'] or 0)
        if r['bulan'] in months:
            idx[key][f"M{r['bulan']}"] += val
            idx[key]["total_sales"] += val

    data_fix = list(idx.values())

    # Footer totals per tahun (hanya dari data tampil)
    footer_totals = {}
    for t in sorted({d["tahun"] for d in data_fix}, reverse=True):
        totals = {f"M{i}": 0 for i in months}
        totals["total_sales"] = 0
        for d in data_fix:
            if d["tahun"] != t:
                continue
            for i in months:
                totals[f"M{i}"] += d[f"M{i}"]
                totals["total_sales"] += d[f"M{i}"]
        footer_totals[t] = totals

    return render_pjax(
        "admin/performa_sales.html",
        data_fix=data_fix,
        list_tahun=list_tahun,
        tahun=thn,
        nama_sales=nama_sales,
        footer_totals=footer_totals,
        filter_bulan=bulan,
        months=months
    )

def parse_amount(x):
    s = str(x or "").replace("Rp", "").replace("rp", "").replace(" ", "").replace(".", "").replace(",", "")
    return int(s) if s.isdigit() else None


def build_date_range(
    year: Optional[int],
    month: Optional[int],
    day: Optional[int] = None,
    alias: str = "bm",        # alias tabel di query
    col: str = "tglfaktur"    # nama kolom tanggal
) -> Tuple[Optional[str], List[date]]:
    """
    Balikkan (clause_sql, params) untuk digunakan di WHERE query:
    "{alias}.{col} >= %s AND {alias}.{col} < %s"

    Contoh:
    - build_date_range(2025, None)           -> tahun penuh 2025
    - build_date_range(None, 5)              -> bulan Mei (semua tahun)
    - build_date_range(None, None, 1)        -> tanggal 1 (semua bulan/tahun)
    - build_date_range(2025, 5)              -> Mei 2025
    - build_date_range(2025, 5, 10)          -> 10 Mei 2025
    """

    clause, params = None, []

    # Tahun + Bulan + Hari
    if year and month and day:
        start = date(year, month, day)
        end = start + timedelta(days=1)
        clause = f"{alias}.{col} >= %s AND {alias}.{col} < %s"
        params = [start, end]

    # Tahun + Bulan
    elif year and month:
        last = calendar.monthrange(year, month)[1]
        start = date(year, month, 1)
        end = date(year, month, last) + timedelta(days=1)
        clause = f"{alias}.{col} >= %s AND {alias}.{col} < %s"
        params = [start, end]

    # Tahun saja
    elif year:
        start = date(year, 1, 1)
        end = date(year + 1, 1, 1)
        clause = f"{alias}.{col} >= %s AND {alias}.{col} < %s"
        params = [start, end]

    # Bulan saja (semua tahun)
    elif month:
        clause = f"MONTH({alias}.{col}) = %s"
        params = [month]

    # Hari saja (semua bulan/tahun)
    elif day:
        clause = f"DAY({alias}.{col}) = %s"
        params = [day]

    return clause, params

@app.route('/admin/performa_sales/export_excel')
def adminperformasalesprint():
    t0 = time.perf_counter()
    log = current_app.logger

    th           = request.args.get('tahun', type=int)
    bln          = request.args.get('bulan', type=int)
    lunas_tidak  = request.args.get('lunas_tidak', type=str)
    nama_sales   = request.args.get('sales')
    raw_angka    = request.args.get('angka')  # opsional; angka minimal
    min_amount   = parse_amount(raw_angka) if 'parse_amount' in globals() else None

    # kolom bulan (1 atau 12)
    months = [bln] if bln else list(range(1, 13))

    # 1) Ambil data agregat sesuai filter (schema baru)
    rows = query_performa(th, bln, nama_sales, lunas_tidak)
    log.debug("[EXPORT] fetched=%d rows; sample_keys=%s",
              len(rows), list(rows[0].keys()) if rows else None)

    # 2) Pivot → baris “gaya Excel”
    pivot = {}
    for r in (rows or []):
        key = (r["tahun"], r["id_sales"], r["nama_outlet"], r["nama_sales"])
        if key not in pivot:
            pivot[key] = {
                "tahun": r["tahun"],
                "id_sales": r["id_sales"],
                "nama_outlet": r["nama_outlet"],
                "nama_sales": r["nama_sales"],
                **{f"M{i}": 0 for i in months},
                "total_sales": 0
            }
        v = int(r.get("nilai") or 0)
        if r["bulan"] in months:
            pivot[key][f"M{r['bulan']}"] += v
            pivot[key]["total_sales"] += v

    data = list(pivot.values())

    # 3) Filter nominal minimal (opsional)
    if min_amount is not None:
        if bln:
            data = [d for d in data if (d.get(f"M{bln}") or 0) >= min_amount]
        else:
            data = [d for d in data if (d.get("total_sales") or 0) >= min_amount]

    # 4) Siapkan workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Performa Sales"

    headers = ["Tahun", "Nama Outlet", "Nama Sales"] + [f"Bulan {m}" for m in months] + ["Jumlah Total"]

    # Judul
    ws["A1"] = "LAPORAN PERFORMA SALES"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")

    # Header tabel (baris 2)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=2, column=c).font = Font(bold=True)

    # 5) Body
    totals = defaultdict(lambda: {"total": 0, **{f"M{i}": 0 for i in months}})

    for item in data:
        vals       = [item["tahun"], item["nama_outlet"], item["nama_sales"]]
        month_vals = [int(item.get(f"M{i}", 0) or 0) for i in months]
        total_val  = int(item.get("total_sales", 0) or 0)

        ws.append(vals + month_vals + [total_val])

        # akumulasi footer per tahun
        for i, v in zip(months, month_vals):
            totals[item["tahun"]][f"M{i}"] += v
            totals[item["tahun"]]["total"] += v

    # 6) Footer JUMLAH per tahun (bold)
    if totals:
        for thn_key in sorted(totals.keys(), reverse=True):
            row = [thn_key, "", "JUMLAH:"] + [totals[thn_key][f"M{i}"] for i in months] + [totals[thn_key]["total"]]
            ws.append(row)
            for c in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    else:
        # kalau tidak ada data, tetap kasih 1 baris kosong agar file valid
        ws.append(["", "", "Tidak ada data untuk filter ini"] + [""] * (len(headers) - 3))

    # 7) Format angka Rp di kolom angka (mulai kolom ke-4)
    #   kolom 1..3 adalah info (tahun/outlet/sales), sisanya angka
    first_amount_col = 4
    last_amount_col  = 3 + len(months) + 1
    for col in range(first_amount_col, last_amount_col + 1):
        for rr in range(3, ws.max_row + 1):  # mulai dari baris 3 (setelah header)
            cell = ws.cell(row=rr, column=col)
            cell.number_format = '"Rp" #,##0'
            cell.alignment = Alignment(horizontal="right")

    # 8) Freeze, filter, width otomatis
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"

    for i in range(1, len(headers) + 1):
        col = get_column_letter(i)
        maxlen = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=i).value
            maxlen = max(maxlen, len(str(v)) if v is not None else 0)
        ws.column_dimensions[col].width = min(50, max(12, maxlen + 2))

    # 9) Output file
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    t1 = time.perf_counter()
    log.info(
        "[EXPORT] excel rows=%d cols=%d duration=%.3fs",
        max(ws.max_row - 2, 0), len(headers), (t1 - t0)
    )

    return send_file(
        bio,
        as_attachment=True,
        download_name="laporan_performa_sales.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/admin/latest_penerimaan', methods=['GET'])
def latest_penerimaan_excell():
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Bulanan"

    bulan_sekarang = time_zone_wib().month
    tahun_ini = time_zone_wib().year
    current_row = 1

    # ===== helper: lebar kolom cepat (tanpa iterasi seluruh sel) =====
    COLS = 11  # A..K
    col_max = [0] * COLS
    def upd_width(vals):
        for i, v in enumerate(vals[:COLS]):
            if v is None:
                continue
            l = len(str(v))
            if l > col_max[i]:
                col_max[i] = l

    # ===== judul umum =====
    ws.merge_cells(f'A{current_row}:K{current_row}')
    ws[f'A{current_row}'] = "LAPORAN BULANAN PERUSAHAAN"
    ws[f'A{current_row}'].font = Font(size=14, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal="center")
    upd_width(["LAPORAN BULANAN PERUSAHAAN"])
    current_row += 2

    # nama bulan lokal
    list_bulan = [
        {"value":"1","nama_bulan":"Januari"},{"value":"2","nama_bulan":"Februari"},
        {"value":"3","nama_bulan":"Maret"},{"value":"4","nama_bulan":"April"},
        {"value":"5","nama_bulan":"Mei"},{"value":"6","nama_bulan":"Juni"},
        {"value":"7","nama_bulan":"Juli"},{"value":"8","nama_bulan":"Agustus"},
        {"value":"9","nama_bulan":"September"},{"value":"10","nama_bulan":"Oktober"},
        {"value":"11","nama_bulan":"November"},{"value":"12","nama_bulan":"Desember"},
    ]
    def nama_bulan_by_num(n:int):
        s = str(n)
        for b in list_bulan:
            if b["value"] == s:
                return b["nama_bulan"]
        return f"Bulan {n}"

    # ===== loop per-bulan =====
    for bulan in range(1, bulan_sekarang + 1):
        # rentang tanggal [start, end)
        last = calendar.monthrange(tahun_ini, bulan)[1]
        start = date(tahun_ini, bulan, 1)
        end   = date(tahun_ini, bulan, last) + timedelta(days=1)

        # --- Ambil data pake skema BARU ---
        #  - purchases (header)
        #  - purchase_items (detail)
        #  - suppliers (nama pemasok)
        #  - products (nama & unit)
        #  - payments (total bayar per invoice; ref_type='PURCHASE')
        hasil = fetch("""
            SELECT
                pu.invoice_date,
                pu.invoice_number,
                sup.name      AS supplier,
                pr.name       AS nama_barang,
                pr.code       AS kode_barang,
                pr.unit       AS unit_produk,
                pi.quantity   AS qty_item,
                pi.unit_price,
                pi.total_amount,
                tot.total_inv,
                COALESCE(pay.total_pay, 0) AS total_pay
            FROM purchases pu
            JOIN purchase_items pi ON pi.purchase_id = pu.id
            JOIN products pr       ON pr.id = pi.product_id
            JOIN suppliers sup     ON sup.id = pu.supplier_id
            LEFT JOIN (
                SELECT purchase_id, SUM(total_amount) AS total_inv
                FROM purchase_items
                GROUP BY purchase_id
            ) tot ON tot.purchase_id = pu.id
            LEFT JOIN (
                SELECT ref_id, SUM(amount) AS total_pay
                FROM payments
                WHERE ref_type = 'PURCHASE'
                GROUP BY ref_id
            ) pay ON pay.ref_id = pu.id
            WHERE pu.invoice_date >= %s AND pu.invoice_date < %s
            ORDER BY pu.invoice_date, pu.invoice_number, pi.id
        """, (start, end))

        if not hasil:
            continue

        # judul bulan
        judul_bulan = f"LAPORAN PEMBELIAN - {nama_bulan_by_num(bulan)} {tahun_ini}"
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = judul_bulan
        ws[f'A{current_row}'].alignment = Alignment(horizontal="center")
        ws[f'A{current_row}'].font = Font(bold=True)
        upd_width([judul_bulan])
        current_row += 1

        # header
        headers = [
            "TANGGAL FAKTUR",   # A
            "NO FAKTUR",        # B
            "SUPPLIER",         # C
            "NAMA BARANG",      # D
            "KODE BARANG",      # E
            "UNIT",             # F (baru)
            "QUANTITY",         # G
            "HARGA SATUAN",     # H
            "HARGA TOTAL",      # I
            "LUNAS / TIDAK",    # J (per-invoice, berdasarkan total pay)
            "PERFORMA BELANJA"  # K (merged block per bulan)
        ]
        ws.append(headers)
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
        upd_width(headers)
        current_row += 1

        performa_belanja = 0
        baris_awal = current_row

        # tulis data baris: status lunas berdasarkan total_pay vs total_inv
        for row in hasil:
            is_lunas = (row.get('total_pay') or 0) >= (row.get('total_inv') or 0)
            if is_lunas:
                # konsisten dengan versi lama: performa = total item yang invoice-nya LUNAS
                performa_belanja += int(row['total_amount'] or 0)

            row_vals = [
                row['invoice_date'].strftime("%Y-%m-%d") if row['invoice_date'] else '',
                row['invoice_number'],
                row['supplier'],
                row['nama_barang'],
                row['kode_barang'],
                row.get('unit_produk') or '',
                row['qty_item'],
                row['unit_price'],
                row['total_amount'],
                "Lunas" if is_lunas else "Tidak Lunas",
                ''  # placeholder kolom performa; nanti di-merge
            ]
            ws.append(row_vals)
            upd_width(row_vals)
            current_row += 1

        # isi & merge kolom performa belanja (K)
        baris_akhir = current_row - 1
        ws.merge_cells(f'K{baris_awal}:K{baris_akhir}')
        ws[f'K{baris_awal}'] = performa_belanja
        ws[f'K{baris_awal}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'K{baris_awal}'].font = Font(bold=True)
        upd_width(['', '', '', '', '', '', '', '', '', '', performa_belanja])

        current_row += 2  # spasi antar bulan

    # format angka Rp untuk kolom H, I, K
    rp_cols = [8, 9, 11]  # 1-based index kolom
    for col in rp_cols:
        for rr in range(3, current_row):
            cell = ws.cell(row=rr, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"Rp" #,##0'
                cell.alignment = Alignment(horizontal="right")

    # set lebar kolom sekali
    for i, mx in enumerate(col_max, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(50, max(12, (mx or 0) + 2))

    # output
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="REPORT_PENERIMAAN.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/admin/latest_pengeluaran', methods=['GET'])
def latest_pengeluaran_excell():
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Bulanan"

    bulan_sekarang = time_zone_wib().month
    tahun_ini = time_zone_wib().year
    current_row = 1

    # ===== helper lebar kolom =====
    COLS = 12
    col_max = [0] * COLS
    def upd_width(vals):
        for i, v in enumerate(vals[:COLS]):
            if v is None: 
                continue
            l = len(str(v))
            if l > col_max[i]:
                col_max[i] = l

    # ===== judul umum =====
    ws.merge_cells(f'A{current_row}:L{current_row}')
    ws[f'A{current_row}'] = "LAPORAN BULANAN PERUSAHAAN"
    ws[f'A{current_row}'].font = Font(size=14, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal="center")
    upd_width(["LAPORAN BULANAN PERUSAHAAN"])
    current_row += 2

    # daftar bulan lokal (hindari dependensi variabel eksternal)
    list_bulan = [
        {"value":"1","nama_bulan":"Januari"},{"value":"2","nama_bulan":"Februari"},
        {"value":"3","nama_bulan":"Maret"},{"value":"4","nama_bulan":"April"},
        {"value":"5","nama_bulan":"Mei"},{"value":"6","nama_bulan":"Juni"},
        {"value":"7","nama_bulan":"Juli"},{"value":"8","nama_bulan":"Agustus"},
        {"value":"9","nama_bulan":"September"},{"value":"10","nama_bulan":"Oktober"},
        {"value":"11","nama_bulan":"November"},{"value":"12","nama_bulan":"Desember"},
    ]
    def nama_bulan_by_num(n:int):
        s = str(n)
        for b in list_bulan:
            if b["value"] == s:
                return b["nama_bulan"]
        return f"Bulan {n}"

    # ===== loop per-bulan =====
    for bulan in range(1, bulan_sekarang + 1):
        # rentang tanggal [start, end)
        last = calendar.monthrange(tahun_ini, bulan)[1]
        start = date(tahun_ini, bulan, 1)
        end   = date(tahun_ini, bulan, last) + timedelta(days=1)

        # Ambil item + meta invoice + total invoice dan total payment (skema BARU)
        hasil = fetch("""
            SELECT
                inv.invoice_date,
                inv.due_date,
                inv.invoice_number,
                sp.name        AS nama_sales,
                c.name         AS nama_outlet,
                p.name         AS nama_barang,
                p.unit         AS unit_produk,
                si.quantity    AS qty_item,
                si.unit_price,
                si.total_amount,
                tot.total_inv,
                COALESCE(pay.total_pay, 0) AS total_pay
            FROM sales_invoices inv
            JOIN sales_items si       ON si.sales_invoice_id = inv.id
            JOIN products p           ON p.id = si.product_id
            JOIN salespersons sp      ON sp.id = inv.salesperson_id
            JOIN customers c          ON c.id = inv.customer_id
            LEFT JOIN (
                SELECT sales_invoice_id, SUM(total_amount) AS total_inv
                FROM sales_items
                GROUP BY sales_invoice_id
            ) tot ON tot.sales_invoice_id = inv.id
            LEFT JOIN (
                SELECT ref_id, SUM(amount) AS total_pay
                FROM payments
                WHERE ref_type='SALE'
                GROUP BY ref_id
            ) pay ON pay.ref_id = inv.id
            WHERE inv.invoice_date >= %s AND inv.invoice_date < %s
            ORDER BY inv.invoice_date, inv.invoice_number, si.id
        """, (start, end))

        if not hasil:
            continue

        # judul bulan
        judul_bulan = f"{nama_bulan_by_num(bulan)} - {tahun_ini}"
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'] = judul_bulan
        ws[f'A{current_row}'].alignment = Alignment(horizontal="center")
        ws[f'A{current_row}'].font = Font(bold=True)
        upd_width([judul_bulan])
        current_row += 1

        # header tabel
        headers = [
            "TANGGAL FAKTUR",      # A
            "JATUH TEMPO",         # B
            "NOMOR FAKTUR",        # C
            "SALES",               # D
            "NAMA OUTLET",         # E
            "NAMA BARANG",         # F
            "UNIT",                # G  (baru: tampilkan unit produk)
            "QUANTITY",            # H
            "HARGA SATUAN",        # I
            "HARGA TOTAL",         # J
            "STATUS (LUNAS?)",     # K
            "PERFORMA SALES"       # L  (merged per bulan)
        ]
        ws.append(headers)
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
        upd_width(headers)
        current_row += 1

        # tulis data + hitung performa_sales (total invoice yang LUNAS di bulan tsb)
        performa_sales = 0
        baris_awal = current_row

        # Kita anggap status LUNAS jika total_pay >= total_inv (per-invoice)
        # Karena hasil berisi banyak baris (per item), status sama untuk semua item dengan invoice yang sama.
        for row in hasil:
            is_lunas = (row.get('total_pay') or 0) >= (row.get('total_inv') or 0)
            if is_lunas:
                # konsisten dengan versi lama: tambahkan total_amount item hanya jika invoice LUNAS
                performa_sales += int(row['total_amount'] or 0)

            row_vals = [
                row['invoice_date'].strftime("%Y-%m-%d") if row['invoice_date'] else '',
                row['due_date'].strftime("%Y-%m-%d") if row['due_date'] else '',
                row['invoice_number'],
                row['nama_sales'],
                row['nama_outlet'],
                row['nama_barang'],
                row.get('unit_produk') or '',
                row['qty_item'],
                row['unit_price'],
                row['total_amount'],
                "Lunas" if is_lunas else "Tidak Lunas",
                ''  # placeholder performa (nanti di-merge)
            ]
            ws.append(row_vals)
            upd_width(row_vals)
            current_row += 1

        # merge & isi kolom performa sales
        baris_akhir = current_row - 1
        ws.merge_cells(f'L{baris_awal}:L{baris_akhir}')
        ws[f'L{baris_awal}'] = performa_sales
        ws[f'L{baris_awal}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'L{baris_awal}'].font = Font(bold=True)
        upd_width(['', '', '', '', '', '', '', '', '', '', '', performa_sales])

        current_row += 2  # spasi antar bulan

    # format angka Rp untuk kolom harga (I, J, L)
    # kolom index: A=1 ... L=12
    rp_cols = [9, 10, 12]
    for col in rp_cols:
        for rr in range(3, current_row):  # mulai dari baris setelah header utama
            cell = ws.cell(row=rr, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"Rp" #,##0'
                cell.alignment = Alignment(horizontal="right")

    # set lebar kolom
    for i, mx in enumerate(col_max, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(50, max(12, (mx or 0) + 2))

    # output excel
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="REPORT_PENGELUARAN.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route("/admin/Report_Laporan")
def report_laporan():
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # ----- Header & Logo -----
    logo_path = "static/image/Logo.png"
    try:
        c.drawImage(logo_path, 1.5*cm, height-4*cm, width=4*cm, preserveAspectRatio=True)
    except Exception:
        pass

    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width/2, height-2*cm, "REPORT LAPORAN BY SYSTEM")

    hari_ini_str = time_zone_wib().strftime("%d-%m-%Y")
    c.setFont("Helvetica", 12)
    c.drawCentredString(width/2, height-2.8*cm, hari_ini_str)

    # ----- Posisi awal tabel -----
    y_pos = height - 5 * cm

    # ====== DATA ======
    today = time_zone_wib().date()
    start_day = today
    end_day   = today + timedelta(days=1)

    data_rows = []

    # INKASO HARIAN (sum payments SALE hari ini)
    g.con.execute("""
        SELECT COALESCE(SUM(p.amount),0)
        FROM payments p
        WHERE p.ref_type='SALE'
          AND p.paid_at >= %s AND p.paid_at < %s
    """, (start_day, end_day))
    inkaso_harian = g.con.fetchone()[0] or 0
    data_rows.append(("INKASO - HARIAN", format_rp(inkaso_harian)))

    # SALES HARIAN (sum sales_items untuk invoice_date = today)
    g.con.execute("""
        SELECT COALESCE(SUM(si.total_amount),0)
        FROM sales_items si
        JOIN sales_invoices inv ON inv.id = si.sales_invoice_id
        WHERE inv.invoice_date = %s
    """, (today,))
    sales_harian = g.con.fetchone()[0] or 0
    data_rows.append(("SALES - HARIAN", format_rp(sales_harian)))

    # Accumulator tahunan
    bulan_sekarang = time_zone_wib().month
    tahun_ini      = time_zone_wib().year

    performa_sales_tahunan   = 0
    performa_belanja_tahunan = 0

    lunas_keuangan  = 0   # SALE paid
    hutang_keuangan = 0   # SALE unpaid

    lunas_admin     = 0   # PURCHASE paid
    piutang_admin   = 0   # PURCHASE unpaid

    # Loop bulanan (1..bulan_sekarang)
    for b in range(1, bulan_sekarang+1):
        # --- SALES (invoice vs payment) ---
        g.con.execute("""
            WITH inv AS (
              SELECT inv.id, SUM(si.total_amount) AS total_inv
              FROM sales_invoices inv
              JOIN sales_items si ON si.sales_invoice_id = inv.id
              WHERE YEAR(inv.invoice_date)=%s AND MONTH(inv.invoice_date)=%s
              GROUP BY inv.id
            ),
            pay AS (
              SELECT ref_id, SUM(amount) AS total_pay
              FROM payments
              WHERE ref_type='SALE'
              GROUP BY ref_id
            )
            SELECT
              COALESCE(SUM(CASE WHEN COALESCE(pay.total_pay,0)>=inv.total_inv THEN inv.total_inv END),0) AS paid,
              COALESCE(SUM(CASE WHEN COALESCE(pay.total_pay,0)< inv.total_inv THEN inv.total_inv END),0) AS unpaid,
              COALESCE(SUM(inv.total_inv),0) AS total_sales
            FROM inv
            LEFT JOIN pay ON pay.ref_id = inv.id
        """, (tahun_ini, b))
        s_paid, s_unpaid, s_total = g.con.fetchone()
        s_paid   = s_paid   or 0
        s_unpaid = s_unpaid or 0
        s_total  = s_total  or 0

        performa_sales_tahunan += s_total
        lunas_keuangan  += s_paid
        hutang_keuangan += s_unpaid

        # --- PURCHASE (invoice vs payment) ---
        g.con.execute("""
            WITH inv AS (
              SELECT pi.id, SUM(pit.total_amount) AS total_inv
              FROM purchase_invoices pi
              JOIN purchase_items pit ON pit.purchase_invoice_id = pi.id
              WHERE YEAR(pi.invoice_date)=%s AND MONTH(pi.invoice_date)=%s
              GROUP BY pi.id
            ),
            pay AS (
              SELECT ref_id, SUM(amount) AS total_pay
              FROM payments
              WHERE ref_type='PURCHASE'
              GROUP BY ref_id
            )
            SELECT
              COALESCE(SUM(CASE WHEN COALESCE(pay.total_pay,0)>=inv.total_inv THEN inv.total_inv END),0) AS paid,
              COALESCE(SUM(CASE WHEN COALESCE(pay.total_pay,0)< inv.total_inv THEN inv.total_inv END),0) AS unpaid,
              COALESCE(SUM(inv.total_inv),0) AS total_purchases
            FROM inv
            LEFT JOIN pay ON pay.ref_id = inv.id
        """, (tahun_ini, b))
        p_paid, p_unpaid, p_total = g.con.fetchone()
        p_paid   = p_paid   or 0
        p_unpaid = p_unpaid or 0
        p_total  = p_total  or 0

        performa_belanja_tahunan += p_total
        lunas_admin  += p_paid
        piutang_admin+= p_unpaid

        # Tambah metrik untuk bulan berjalan saja
        if b == bulan_sekarang:
            data_rows.append(("PERFORMA SALES - BULANAN",   format_rp(s_total)))
            data_rows.append(("PERFORMA BELANJA - BULANAN", format_rp(p_total)))
            data_rows.append(("CASHFLOW - BULANAN",         format_rp(s_total - p_total)))

    # Rekap tahunan
    data_rows.append(("PERFORMA SALES - TAHUNAN",    format_rp(performa_sales_tahunan)))
    data_rows.append(("PERFORMA BELANJA - TAHUNAN",  format_rp(performa_belanja_tahunan)))
    data_rows.append(("CASHFLOW - TAHUNAN",          format_rp(performa_sales_tahunan - performa_belanja_tahunan)))
    data_rows.append(("LUNAS - KEUANGAN (SALE)",     format_rp(lunas_keuangan)))
    data_rows.append(("TIDAK LUNAS - KEUANGAN",      format_rp(hutang_keuangan)))
    data_rows.append(("LUNAS - ADMINISTRASI (PUR)",  format_rp(lunas_admin)))
    data_rows.append(("TIDAK LUNAS - ADMINISTRASI",  format_rp(piutang_admin)))

    # ====== RENDER TABEL ======
    # siapkan data untuk Table: kolom No, Keterangan, Nilai
    table_data = [["No", "Keterangan", "Nilai (Rp)"]]
    for i, (ket, nilai) in enumerate(data_rows, start=1):
        table_data.append([str(i), ket, str(nilai)])

    # hitung lebar kolom proporsional
    # jaga agar tabel muat halaman (margin kiri=2cm, kanan=2cm)
    max_table_width = width - 4*cm
    col3_min = stringWidth("Rp 0", "Helvetica", 10) + 40
    col1_w = 1.6*cm
    col3_w = max(col3_min, 4*cm)
    col2_w = max_table_width - col1_w - col3_w
    if col2_w < 6*cm:  # fallback kalau terlalu sempit
        col2_w = 6*cm
        col3_w = max_table_width - col1_w - col2_w

    tbl = Table(table_data, colWidths=[col1_w, col2_w, col3_w])
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("ALIGN", (0,0), (0,-1), "CENTER"),   # kolom No center
        ("ALIGN", (1,0), (1,-1), "LEFT"),     # kolom Keterangan kiri
        ("ALIGN", (2,0), (2,-1), "RIGHT"),    # kolom Nilai kanan
        ("BACKGROUND", (0,0), (-1,0), colors.whitesmoke),
        ("TEXTCOLOR", (0,0), (-1,0), colors.black),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 4),
    ]))

    # Auto page-break sederhana: pindah ke halaman baru jika tingginya turun <2cm
    # (ReportLab Table nggak auto flow, jadi kita render sekali. Kalau yakin datanya panjang,
    # pindahkan ke SimpleDocTemplate + Flowables. Untuk laporan ringkas ini cukup.)
    tbl_h = 0.6*cm * len(table_data) + 1*cm  # estimasi kasar
    if y_pos - tbl_h < 2*cm:
        c.showPage()
        y_pos = height - 3*cm

    tbl.wrapOn(c, width, height)
    tbl.drawOn(c, 2*cm, y_pos - tbl_h + 0.4*cm)

    c.save()
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name="REPORT_LAPORAN.pdf",
                     mimetype="application/pdf")